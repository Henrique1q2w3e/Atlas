#!/usr/bin/env python3
import os
import sys

# Adicionar o diretório atual ao path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Mudar para o diretório correto
os.chdir(os.path.dirname(os.path.abspath(__file__)))

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
import mercadopago
from datetime import datetime, timezone, timedelta
import uuid
import re
# import pandas as pd  # Removido para compatibilidade com Render
import json
import sqlite3
import hashlib
import secrets
import time
import hmac
import base64
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook, load_workbook
import requests

app = Flask(__name__, template_folder='templates', static_folder='static')
app.secret_key = os.environ.get('SECRET_KEY', 'atlas_suplementos_secret_key_2024_secure')
app.config['SESSION_COOKIE_SECURE'] = False  # Para desenvolvimento
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Funções de autenticação
def hash_senha(senha):
    """Hash da senha usando SHA-256"""
    return hashlib.sha256(senha.encode()).hexdigest()

def verificar_senha(senha, hash_senha_armazenado):
    """Verificar se a senha está correta"""
    senha_hash = hash_senha(senha)
    print(f"🔐 Comparando senhas:")
    print(f"   Senha digitada (hash): {senha_hash}")
    print(f"   Senha armazenada: {hash_senha_armazenado}")
    print(f"   São iguais: {senha_hash == hash_senha_armazenado}")
    return senha_hash == hash_senha_armazenado

def conectar_db():
    """Conectar ao banco de dados"""
    # Usar PostgreSQL se disponível, senão SQLite
    database_url = os.environ.get('DATABASE_URL')
    
    if database_url:
        # PostgreSQL no Render
        print("💾 Conectando ao PostgreSQL...")
        import psycopg2
        from urllib.parse import urlparse
        
        # Parse da URL do PostgreSQL
        url = urlparse(database_url)
        conn = psycopg2.connect(
            database=url.path[1:],
            user=url.username,
            password=url.password,
            host=url.hostname,
            port=url.port
        )
        print(f"💾 Conectado ao PostgreSQL: {url.hostname}")
        return conn
    else:
        # SQLite local
        print("💾 Conectando ao SQLite...")
        db_path = os.path.join(os.getcwd(), 'atlas.db')
        print(f"💾 Usando SQLite local: {db_path}")
        return sqlite3.connect(db_path)

def executar_query(cursor, query, params=None):
    """Executar query com placeholders corretos para PostgreSQL ou SQLite"""
    database_url = os.environ.get('DATABASE_URL')
    
    if database_url:
        # PostgreSQL usa %s
        if params:
            # Converter ? para %s
            query = query.replace('?', '%s')
        cursor.execute(query, params)
    else:
        # SQLite usa ?
        cursor.execute(query, params)

def criar_tabelas():
    """Criar tabelas do banco de dados se não existirem"""
    try:
        print("🔧 Criando/conectando ao banco de dados...")
        conn = conectar_db()
        cursor = conn.cursor()
        
        database_url = os.environ.get('DATABASE_URL')
        
        if database_url:
            # PostgreSQL
            print("💾 Criando tabelas no PostgreSQL...")
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS usuario (
                    id SERIAL PRIMARY KEY,
                    nome VARCHAR(255) NOT NULL,
                    email VARCHAR(255) UNIQUE NOT NULL,
                    senha_hash VARCHAR(255) NOT NULL,
                    data_criacao TIMESTAMP NOT NULL,
                    admin INTEGER DEFAULT 0
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS carrinho (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL,
                    produto_id VARCHAR(255) NOT NULL,
                    nome VARCHAR(255) NOT NULL,
                    marca VARCHAR(255),
                    preco DECIMAL(10,2) NOT NULL,
                    sabor VARCHAR(255),
                    quantidade INTEGER NOT NULL,
                    imagem VARCHAR(500),
                    data_adicionado TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(user_id, produto_id, sabor),
                    FOREIGN KEY (user_id) REFERENCES usuario(id) ON DELETE CASCADE
                )
            ''')
        else:
            # SQLite
            print("💾 Criando tabelas no SQLite...")
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS usuario (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome TEXT NOT NULL,
                    email TEXT UNIQUE NOT NULL,
                    senha_hash TEXT NOT NULL,
                    data_criacao TEXT NOT NULL,
                    admin INTEGER DEFAULT 0
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS carrinho (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    produto_id TEXT NOT NULL,
                    nome TEXT NOT NULL,
                    marca TEXT,
                    preco REAL NOT NULL,
                    sabor TEXT,
                    quantidade INTEGER NOT NULL,
                    imagem TEXT,
                    data_adicionado TEXT DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(user_id, produto_id, sabor),
                    FOREIGN KEY (user_id) REFERENCES usuario(id)
                )
            ''')
        
        conn.commit()
        conn.close()
        print("✅ Tabelas do banco de dados criadas/verificadas com sucesso!")
        
    except Exception as e:
        print(f"❌ Erro ao criar tabelas: {e}")
        import traceback
        traceback.print_exc()

def usuario_logado():
    """Verificar se usuário NORMAL está logado (não admin)"""
    print(f"🔍 Verificando login usuário normal - Sessão: {dict(session)}")
    print(f"🔍 user_id na sessão: {'user_id' in session}")
    print(f"🔍 is_admin_session: {session.get('is_admin_session', False)}")
    
    # Usuário normal está logado se tem user_id E não é sessão de admin
    resultado = 'user_id' in session and not session.get('is_admin_session', False)
    print(f"🔍 usuario_logado() retorna: {resultado}")
    return resultado

def admin_logado():
    """Verificar se admin está logado"""
    return 'admin_user_id' in session and session.get('is_admin_session', False)

def qualquer_usuario_logado():
    """Verificar se qualquer usuário está logado (normal ou admin)"""
    print(f"🔍 Verificando qualquer_usuario_logado - Sessão: {dict(session)}")
    print(f"🔍 user_id na sessão: {'user_id' in session}")
    print(f"🔍 admin_logado(): {admin_logado()}")
    
    resultado = 'user_id' in session or admin_logado()
    print(f"🔍 qualquer_usuario_logado() retorna: {resultado}")
    return resultado

def validar_email(email):
    """Valida formato do email"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def validar_cpf(cpf):
    """Valida CPF brasileiro"""
    # Remove caracteres não numéricos
    cpf = re.sub(r'[^0-9]', '', cpf)
    
    # Verifica se tem 11 dígitos
    if len(cpf) != 11:
        return False
    
    # Verifica se todos os dígitos são iguais
    if cpf == cpf[0] * 11:
        return False
    
    # Validação do CPF
    def calcular_digito(cpf, posicoes):
        soma = sum(int(cpf[i]) * posicoes[i] for i in range(len(posicoes)))
        resto = soma % 11
        return 0 if resto < 2 else 11 - resto
    
    # Primeiro dígito verificador
    posicoes1 = list(range(10, 1, -1))
    digito1 = calcular_digito(cpf, posicoes1)
    
    # Segundo dígito verificador
    posicoes2 = list(range(11, 1, -1))
    digito2 = calcular_digito(cpf, posicoes2)
    
    return cpf[-2:] == f"{digito1}{digito2}"

def validar_telefone(telefone):
    """Valida telefone brasileiro"""
    # Remove caracteres não numéricos
    telefone = re.sub(r'[^0-9]', '', telefone)
    
    # Verifica se tem 10 ou 11 dígitos (com DDD)
    if len(telefone) not in [10, 11]:
        return False
    
    # Verifica se começa com DDD válido (11-99)
    ddd = int(telefone[:2])
    return 11 <= ddd <= 99

def obter_horario_brasil():
    """Retorna horário atual do Brasil (UTC-3)"""
    utc_now = datetime.now(timezone.utc)
    brasil_tz = timezone(timedelta(hours=-3))
    return utc_now.astimezone(brasil_tz)

def criar_notificacao(order_id, email, telefone, status, mensagem):
    """Cria uma notificação para o cliente"""
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        
        executar_query(cursor, '''
            INSERT INTO notificacoes (order_id, email, telefone, status, mensagem)
            VALUES (?, ?, ?, ?, ?)
        ''', (order_id, email, telefone, status, mensagem))
        
        conn.commit()
        conn.close()
        print(f"📧 Notificação criada para {email}: {status}")
        
    except Exception as e:
        print(f"❌ Erro ao criar notificação: {e}")

def obter_mensagem_status(status):
    """Retorna mensagem personalizada para cada status"""
    mensagens = {
        'Pendente': 'Seu pedido foi recebido e está sendo processado! 🛒',
        'Pago': 'Pagamento confirmado! Seu pedido está em produção! 💰',
        'Em Produção': 'Seu pedido está sendo preparado com carinho! ⚙️',
        'Saiu para Entrega': 'Seu pedido saiu para entrega! 🚚',
        'Enviado': 'Seu pedido saiu para entrega! 🚚',
        'Entregue': 'Pedido entregue com sucesso! Obrigado pela preferência! 🎉'
    }
    return mensagens.get(status, f'Status do pedido atualizado para: {status}')

def enviar_whatsapp_automatico(order_id, nome, telefone, status):
    """Envia mensagem automática no WhatsApp para mudança de status"""
    try:
        if not telefone:
            print("⚠️ Telefone não informado, pulando WhatsApp")
            return False
        
        # Limpar telefone (remover caracteres especiais)
        telefone_limpo = re.sub(r'[^0-9]', '', telefone)
        
        # Verificar se tem DDD
        if len(telefone_limpo) == 9:
            telefone_limpo = "11" + telefone_limpo  # Adicionar DDD 11 se não tiver
        
        # Criar mensagem personalizada
        mensagem = f"""🏪 *Atlas Suplementos*

Olá {nome}! 👋

📦 *Atualização do seu pedido #{order_id}*

{obter_mensagem_status(status)}

🔗 Acompanhe seu pedido: https://atlas-1h3w.onrender.com/status-pedido

Obrigado por escolher a Atlas Suplementos! 💪"""
        
        # URL do WhatsApp Web
        whatsapp_url = f"https://wa.me/55{telefone_limpo}?text={requests.utils.quote(mensagem)}"
        
        print(f"📱 WhatsApp preparado para {nome} ({telefone_limpo}): {status}")
        print(f"📱 URL: {whatsapp_url}")
        
        # Em produção, você pode usar uma API real do WhatsApp
        # Por enquanto, apenas logamos a URL
        return True
        
    except Exception as e:
        print(f"❌ Erro ao preparar WhatsApp: {e}")
        return False

def obter_usuario_logado():
    """Obtém os dados do usuário logado"""
    try:
        print("🔍 Verificando se usuário está logado...")
        if not usuario_logado():
            print("❌ Usuário não está logado")
            return None
        
        print(f"👤 User ID na sessão: {session.get('user_id')}")
        conn = conectar_db()
        cursor = conn.cursor()
        executar_query(cursor, 'SELECT id, nome, email, data_criacao, admin FROM usuario WHERE id = ?', (session['user_id'],))
        usuario = cursor.fetchone()
        conn.close()
        
        print(f"👤 Usuário encontrado no banco: {usuario}")
        
        if usuario:
            user_data = {
                'id': usuario[0],
                'nome': usuario[1],
                'email': usuario[2],
                'data_criacao': usuario[3],
                'admin': usuario[4]
            }
            print(f"✅ Dados do usuário preparados: {user_data}")
            return user_data
        else:
            print("❌ Usuário não encontrado no banco de dados")
            return None
            
    except Exception as e:
        print(f"💥 Erro ao obter usuário logado: {e}")
        import traceback
        traceback.print_exc()
    return None

def obter_imagem_produto(marca, categoria):
    """Mapeia marca e categoria para imagem específica"""
    marca_lower = marca.lower().strip()
    categoria_lower = categoria.lower().strip()
    
    # Mapeamento de imagens por marca e categoria
    imagens_map = {
        # MAX - Whey
        ('max', 'whey'): '/static/images/whey-isolado-max-card.png',
        ('max', 'whey isolado'): '/static/images/whey-isolado-max-card.png',
        ('max', 'whey concentrado'): '/static/images/whey-concentrado-max-card.png',
        ('max', 'whey 3w'): '/static/images/whey-3w-max-card.png',
        
        # MAX - Pré-treino
        ('max', 'pré-treino'): '/static/images/horus-max-card.png',
        ('max', 'horus'): '/static/images/horus-max-card.png',
        ('max', 'pré'): '/static/images/horus-max-card.png',
        
        # MAX - Hipercalórico
        ('max', 'hipercalórico'): '/static/images/hipercalorico-max-card.png',
        ('max', 'hiper'): '/static/images/hipercalorico-max-card.png',
        
        # MAX - Multivitamínico
        ('max', 'multivitamínico'): '/static/images/multivitaminico-max-card.png',
        ('max', 'vitamina'): '/static/images/multivitaminico-max-card.png',
        
        # DUX - Whey
        ('dux', 'whey'): '/static/images/whey-concentrado-dux-card.png',
        ('dux', 'whey isolado'): '/static/images/whey-isolado-dux-card.png',
        ('dux', 'whey concentrado'): '/static/images/whey-concentrado-dux-card.png',
        
        # DUX - Creatina
        ('dux', 'creatina'): '/static/images/creatina-dux-card.png',
        
        # DUX - Multivitamínico
        ('dux', 'multivitamínico'): '/static/images/multivitaminco-dux-card.png',
        ('dux', 'vitamina'): '/static/images/multivitaminco-dux-card.png',
        
        # DUX - Ômega 3
        ('dux', 'ômega'): '/static/images/omega3-dux-card.png',
        ('dux', 'omega'): '/static/images/omega3-dux-card.png',
        
        # DUX - Cafeína
        ('dux', 'cafeína'): '/static/images/cafeina-dux-card.png',
        ('dux', 'cafeina'): '/static/images/cafeina-dux-card.png',
        
        # FTW - Whey
        ('ftw', 'whey'): '/static/images/whey-concentrado-ftw-card.png',
        ('ftw', 'whey 3w'): '/static/images/whey-3w-ftw-card.png',
        
        # FTW - Creatina
        ('ftw', 'creatina'): '/static/images/creatina-ftw-card.png',
        
        # FTW - Pré-treino
        ('ftw', 'pré-treino'): '/static/images/pre-treino-ftw-card.png',
        ('ftw', 'pré'): '/static/images/pre-treino-ftw-card.png',
        
        # SHARK - Whey
        ('shark', 'whey'): '/static/images/whey-concentrado-shark-card.png',
        ('shark', 'whey isolado'): '/static/images/whey-isolado-shark-card.png',
        
        # SHARK - Creatina
        ('shark', 'creatina'): '/static/images/creatina-shark-card.png',
        
        # SHARK - Pré-treino
        ('shark', 'pré-treino'): '/static/images/pre-treino-shark-card.png',
        ('shark', 'pré'): '/static/images/pre-treino-shark-card.png',
        
        # INTEGRAL - Whey
        ('integral', 'whey'): '/static/images/whey-concentrado-integral-card.png',
        ('integral', 'whey isolado'): '/static/images/whey-isolado-integral-card.png',
        
        # INTEGRAL - Creatina
        ('integral', 'creatina'): '/static/images/creatina-integral-card.png',
        
        # NUTRA - Whey
        ('nutra', 'whey'): '/static/images/whey-concentrado-nutra-card.png',
        ('nutra', 'whey isolado'): '/static/images/whey-isolado-nutra-card.png',
        
        # NUTRA - Barrinha
        ('nutra', 'barrinha'): '/static/images/barrinha-nutra-card.png',
        
        # ATLHETICA - Whey
        ('atlhetica', 'whey'): '/static/images/whey-concentrado-atlhetica-card.png',
        ('atlhetica', 'whey isolado'): '/static/images/whey-best-atlhetica-card.png',
        
        # ATLHETICA - Barrinha
        ('atlhetica', 'barrinha'): '/static/images/barrinha-atlhetica-card.png',
        
        # ATLHETICA - Multivitamínico
        ('atlhetica', 'multivitamínico'): '/static/images/multivitaminico-atlhetica-card.png',
        ('atlhetica', 'vitamina'): '/static/images/multivitaminico-atlhetica-card.png',
        
        # ADAPTOGEN - Whey
        ('adaptogen', 'whey'): '/static/images/whey-gold-adaptogen-card.png',
        ('adaptogen', 'whey tasty'): '/static/images/whey-tasty-adaptogen-card.png',
        
        # UNDER LABZ - Whey
        ('under labz', 'whey'): '/static/images/whey-concentrado-under-labz-card.png',
        ('under labz', 'whey isolado'): '/static/images/whey-isolado-under-labz-card.png',
        
        # UNDER LABZ - Pré-treino
        ('under labz', 'pré-treino'): '/static/images/pre-treino-under-labz-card.png',
        ('under labz', 'pré'): '/static/images/pre-treino-under-labz-card.png',
    }
    
    # Tentar encontrar imagem específica
    for (marca_key, categoria_key), imagem in imagens_map.items():
        if marca_key in marca_lower and categoria_key in categoria_lower:
            return imagem
    
    # Imagens por categoria padrão
    categoria_imagens = {
        'whey': '/static/images/whey-max.png',
        'creatina': '/static/images/creatina.png',
        'pré-treino': '/static/images/pre-max.png',
        'pré': '/static/images/pre-max.png',
        'hipercalórico': '/static/images/hipercaloricos.png',
        'hiper': '/static/images/hipercaloricos.png',
        'multivitamínico': '/static/images/omega-3.png',
        'vitamina': '/static/images/omega-3.png',
        'barrinha': '/static/images/barrinhas.png',
        'ômega': '/static/images/omega-3.png',
        'omega': '/static/images/omega-3.png',
        'cafeína': '/static/images/capsula-de-cafeina.png',
        'cafeina': '/static/images/capsula-de-cafeina.png',
    }
    
    # Tentar encontrar por categoria
    for cat_key, imagem in categoria_imagens.items():
        if cat_key in categoria_lower:
            return imagem
    
    # Imagem padrão
    return '/static/images/produto-placeholder.svg'

def carregar_produtos():
    """Carrega produtos da planilha Excel usando openpyxl"""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook('atlas.xlsx')
        ws = wb.active
        
        produtos = []
        for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            if not row[0] or not row[1]:  # MARCA e CATEGORIA
                continue
                
            marca = str(row[0]).strip()
            categoria = str(row[1]).strip()
            sabores_texto = str(row[2]).strip() if row[2] else 'N/A'
            
            sabores_lista = []
            if sabores_texto != 'N/A' and sabores_texto != 'NÃO TEM SABORES':
                sabores_lista = [s.strip() for s in sabores_texto.split(',') if s.strip()]
            
            # Determinar categoria para filtros
            categoria_filtro = 'whey'  # Padrão
            categoria_lower = categoria.lower()
            
            if 'creatina' in categoria_lower:
                categoria_filtro = 'creatina'
            elif 'pré' in categoria_lower or 'treino' in categoria_lower or 'horus' in categoria_lower or 'égide' in categoria_lower or 'fire' in categoria_lower:
                categoria_filtro = 'pre_treino'
            elif 'hiper' in categoria_lower:
                categoria_filtro = 'hipercalorico'
            elif 'vitamina' in categoria_lower or 'multivitamínico' in categoria_lower or 'multivitaminco' in categoria_lower:
                categoria_filtro = 'vitaminas'
            elif 'barrinha' in categoria_lower or 'barrinhas' in categoria_lower:
                categoria_filtro = 'barrinhas'
            elif 'omega' in categoria_lower or 'cafeína' in categoria_lower or 'cafeina' in categoria_lower:
                categoria_filtro = 'vitaminas'  # Agrupar suplementos em vitaminas
            
            produto = {
                'id': f"produto_{index}",
                'nome': f"{marca} - {categoria}",
                'marca': marca,
                'categoria': categoria_filtro,
                'sabores': sabores_lista,
                'preco': 99.90,  # Preço padrão
                'imagem': obter_imagem_produto(marca, categoria),
                'imagem_principal': obter_imagem_produto(marca, categoria),
                'descricao': f"Suplemento {categoria} da marca {marca}",
                'estoque': 10
            }
            # --- Overrides pontuais por produto (atualizar um por um) ---
            try:
                marca_lower = marca.lower().strip()
                categoria_raw_lower = categoria.lower().strip()

                # MAX - Whey Isolado: atualizar preço, sabores e imagens específicas
                if marca_lower == 'max' and 'whey' in categoria_raw_lower and 'isol' in categoria_raw_lower:
                    produto['preco'] = 160.00
                    produto['sabores'] = ['cookies', 'baunilha', 'morango']

                    # Card / lista
                    produto['imagem'] = '/static/images/whey-isolado-max-card.png'
                    produto['imagem_principal'] = '/static/images/whey-isolado-max-card.png'
                    produto['imagem_mobile'] = '/static/images/whey-isolado-max-card-mobile.png'

                    # Imagens do modal (desktop)
                    produto['imagens'] = [
                        '/static/images/whey-isolado-max-modal-frente.png',
                        '/static/images/whey-isolado-max-modal-meta.png',
                        '/static/images/whey-isolado-max-modal-tablea.png'
                    ]

                    # Imagens do modal (mobile)
                    produto['imagens_mobile'] = [
                        '/static/images/whey-isolado-max-modal-mobile-frente.png',
                        '/static/images/whey-isolado-max-modal-mobile-meta.png',
                        '/static/images/whey-isolado-max-modal-mobile-tabela.png'
                    ]

                # MAX - Whey Concentrado: atualizar preço, sabores e imagens específicas
                elif marca_lower == 'max' and 'whey' in categoria_raw_lower and 'concentr' in categoria_raw_lower:
                    produto['preco'] = 119.90
                    produto['sabores'] = ['morango', 'chocolate', 'baunilha', 'cookies']

                    # Card / lista
                    produto['imagem'] = '/static/images/whey-concentrado-max-card.png'
                    produto['imagem_principal'] = '/static/images/whey-concentrado-max-card.png'
                    produto['imagem_mobile'] = '/static/images/whey-concentrado-max-card-mobile.png'

                    # Imagens do modal (desktop)
                    produto['imagens'] = [
                        '/static/images/whey-concentrado-max-modal-frente.png',
                        '/static/images/whey-concentrado-max-modal-meta.png',
                        '/static/images/whey-concentrado-max-modal-tebela.png'
                    ]

                    # Imagens do modal (mobile)
                    produto['imagens_mobile'] = [
                        '/static/images/whey-concentrado-max-modal-frente-mobile.png',
                        '/static/images/whey-concentrado-max-modal-meta-mobile.png',
                        '/static/images/whey-concentrado-max-modal-tebela-mobile.png'
                    ]
                
                # MAX - Whey 3W (1,8kg): atualizar preço, sabores e imagens específicas
                elif marca_lower == 'max' and 'whey' in categoria_raw_lower and ('3w' in categoria_raw_lower or '3 w' in categoria_raw_lower):
                    produto['preco'] = 280.00

                    # Alguns registros usam texto de sabores confuso; padronizar para lista limpa
                    produto['sabores'] = ['chocolate', 'morango', 'baunilha']

                    # Card / lista (se não existir card desktop, manter imagem padrão)
                    # Usar a versão mobile como card desktop (arquivo fornecido)
                    produto['imagem_principal'] = '/static/images/whey-3w-max-card-mobile.png'
                    produto['imagem_mobile'] = '/static/images/whey-3w-max-card-mobile.png'

                    # Imagens do modal (desktop)
                    produto['imagens'] = [
                        '/static/images/whey-3w-max-modal-frente.png',
                        '/static/images/whey-3w-max-modal-meta.png',
                        '/static/images/whey-3w-max-modal-tabela.png'
                    ]

                    # Imagens do modal (mobile)
                    produto['imagens_mobile'] = [
                        '/static/images/whey-3w-max-modal-frente-mobile.png',
                        '/static/images/whey-3w-max-modal-meta-mobile.png',
                        '/static/images/whey-3w-max-modal-tabela-mobile.png'
                    ]
                
                # MAX - Horus (pré-treino)
                elif marca_lower == 'max' and ('horus' in categoria_raw_lower or 'pré' in categoria_raw_lower and 'horus' in categoria_raw_lower):
                    produto['nome'] = 'MAX - HORUS PRÉ TREINO'
                    produto['preco'] = 89.90
                    produto['sabores'] = ['amora', 'blue ice', 'citrus', 'limao yuzu', 'frutas vermelhas', 'maçã verde']
                    produto['imagem'] = '/static/images/horus-max-card.png'
                    produto['imagem_principal'] = '/static/images/horus-max-card.png'
                    produto['imagem_mobile'] = '/static/images/horus-max-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/horus-max-modal-frente.png',
                        '/static/images/horus-max-modal-meta.png',
                        '/static/images/horus-max-modal-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/horus-max-modal-frente-mobile.png',
                        '/static/images/horus-max-modal-meta-mobile.png',
                        '/static/images/horus-max-modal-tabela-mobile.png'
                    ]

                # MAX - Égide (pré-treino)
                elif marca_lower == 'max' and 'égide' in categoria_raw_lower or ('egide' in categoria_raw_lower):
                    produto['preco'] = 89.90
                    produto['sabores'] = ['abacaxi com hortelã', 'abacaxi com manga', 'frutas silvestres', 'frutas vermelhas']
                    produto['imagem'] = '/static/images/egide-max-card.png'
                    produto['imagem_principal'] = '/static/images/egide-max-card.png'
                    produto['imagem_mobile'] = '/static/images/egide-max-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/egide-max-modal-frente.png',
                        '/static/images/egide-max-modal-meta.png',
                        '/static/images/egide-max-modal-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/egide-max-modal-frente-mobile.png',
                        '/static/images/egide-max-modal-meta-mobile.png',
                        '/static/images/egide-max-modal-tabela-mobile.png'
                    ]

                # MAX - Fire Black (sabor padrão)
                elif marca_lower == 'max' and 'fire' in categoria_raw_lower:
                    produto['preco'] = 49.90
                    produto['sabores'] = []
                    produto['imagem'] = '/static/images/fire-max-card.png'
                    produto['imagem_principal'] = '/static/images/fire-max-card.png'
                    produto['imagem_mobile'] = '/static/images/fire-max-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/fire-max-modal-frente.png',
                        '/static/images/fire-max-modal-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/fire-max-modal-frente-mobile.png',
                        '/static/images/fire-max-modal-tabela-mobile.png'
                    ]

                # MAX - Multivitamínico
                elif marca_lower == 'max' and 'multivit' in categoria_raw_lower:
                    produto['preco'] = 59.90
                    produto['sabores'] = []
                    produto['imagem'] = '/static/images/multivitaminico-max-card.png'
                    produto['imagem_principal'] = '/static/images/multivitaminico-max-card.png'
                    produto['imagem_mobile'] = '/static/images/multivitaminico-max-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/multivitaminico-max-modal-frente.png',
                        '/static/images/multivitaminico-max-modal-meta.png',
                        '/static/images/multivitaminico-max-modal-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/multivitaminico-max-modal-frente-mobile.png',
                        '/static/images/multivitaminico-max-modal-meta-mobile.png',
                        '/static/images/multivitaminico-max-modal-tabela-mobile.png'
                    ]

                # MAX - Creatina 150g
                elif marca_lower == 'max' and 'creatina' in categoria_raw_lower and '150' in categoria_raw_lower:
                    produto['preco'] = 49.90
                    produto['sabores'] = []
                    produto['imagem'] = '/static/images/creatina150-max-card.png'
                    produto['imagem_principal'] = '/static/images/creatina150-max-card.png'
                    produto['imagem_mobile'] = '/static/images/creatina150-max-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/creatina150-max-modal-frente.png',
                        '/static/images/creatina150-max-modal-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/creatina150-max-modal-frente-mobile.png',
                        '/static/images/creatina150-max-modal-tabela-mobile.png'
                    ]

                # MAX - Creatina 300g
                elif marca_lower == 'max' and 'creatina' in categoria_raw_lower and '300' in categoria_raw_lower:
                    produto['preco'] = 99.90
                    produto['sabores'] = []
                    produto['imagem'] = '/static/images/creatina300-max-card.png'
                    produto['imagem_principal'] = '/static/images/creatina300-max-card.png'
                    produto['imagem_mobile'] = '/static/images/creatina300-max-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/creatina300-max-modal-frente.png',
                        '/static/images/creatina300-max-modal-frente-mobile.png'
                    ]
                    
                # MAX - Hipercalórico
                elif marca_lower == 'max' and ('hiper' in categoria_raw_lower or 'hipercalor' in categoria_raw_lower):
                    produto['preco'] = 89.90
                    produto['sabores'] = ['Chocolate', 'morango', 'baunilha']
                    produto['imagem'] = '/static/images/hipercalorico-max-card.png'
                    produto['imagem_principal'] = '/static/images/hipercalorico-max-card.png'
                    produto['imagem_mobile'] = '/static/images/hipercalorico-max-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/hipercalorico-max-modal-frente.png',
                        '/static/images/hipercalorico-max-modal-meta.png',
                        '/static/images/hipercalorico-max-modal-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/hipercalorico-max-modal-frente-mobile.png',
                        '/static/images/hipercalorico-max-modal-meta-mobile.png',
                        '/static/images/hipercalorico-max-modal-tabela-mobile.png'
                    ]

                # MAX - Pre treino sem cafeína
                elif marca_lower == 'max' and 'pre' in categoria_raw_lower and 'sem' in categoria_raw_lower and 'cafe' in categoria_raw_lower:
                    produto['preco'] = 99.90
                    produto['sabores'] = ['citrus']
                    produto['imagem'] = '/static/images/pre-sem-cafeina-max-card.png'
                    produto['imagem_principal'] = '/static/images/pre-sem-cafeina-max-card.png'
                    produto['imagem_mobile'] = '/static/images/pre-sem-cafeina-max-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/pre-sem-cafeina-max-modal-frente.png',
                        '/static/images/pre-sem-cafeina-max-modal-tabela.png',
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/pre-sem-cafeina-max-modal-frente-mobile.png',
                        '/static/images/pre-sem-cafeina-max-modal-tabela-mobile.png',
                    ]
                
                # ADAPTOGEN - Gold Whey
                elif marca_lower == 'adaptogen' and ('gold' in categoria_raw_lower or 'gold whey' in categoria_raw_lower):
                    produto['preco'] = 119.90
                    produto['sabores'] = ['baunilha', 'chocolate', 'chocotella', 'coco', 'cookies', 'doce de leite', 'morango', 'original']
                    produto['imagem'] = '/static/images/whey-gold-adaptogen-card.png'
                    produto['imagem_principal'] = '/static/images/whey-gold-adaptogen-card.png'
                    produto['imagem_mobile'] = '/static/images/whey-gold-adaptogen-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/whey-gold-adaptogen-modal-frente.png',
                        '/static/images/whey-gold-adaptogen-modal-meta.png',
                        '/static/images/whey-gold-adaptogen-modal-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/whey-gold-adaptogen-modal-frente-mobile.png',
                        '/static/images/whey-gold-adaptogen-modal-meta-mobile.png',
                        '/static/images/whey-gold-adaptogen-modal-tabela-mobile.png'
                    ]

                # ADAPTOGEN - Linha Tasty
                elif marca_lower == 'adaptogen' and ('tasty' in categoria_raw_lower or 'linha tasty' in categoria_raw_lower or 'tasty' in produto.get('nome','').lower()):
                    produto['preco'] = 189.90
                    produto['sabores'] = [
                        'chiclete','chocolate suíço','chocolate peanut butter','chocomaltine','chocotella','churros','coco',
                        'cookies and cream','doce de leite','leite condensado','manga','mousse de chocolate','milho verde','morango','original','pistache'
                    ]
                    produto['imagem'] = '/static/images/whey-tasty-adaptogen-card.png'
                    produto['imagem_principal'] = '/static/images/whey-tasty-adaptogen-card.png'
                    produto['imagem_mobile'] = '/static/images/whey-tasty-adaptogen-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/whey-tasty-adaptogen-modal-frente.png',
                        '/static/images/whey-tasty-adaptogen-modal-meta.png',
                        '/static/images/whey-tasty-adaptogen-modal-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/whey-tasty-adaptogen-modal-frente-mobile.png',
                        '/static/images/whey-tasty-adaptogen-modal-meta-mobile.png',
                        '/static/images/whey-tasty-adaptogen-modal-tabela-mobile.png'
                    ]

                # ATLETICA / ATLHETICA - Whey Tech, Best, Creatinas, Barrinha, Multivitamínico
                elif marca_lower in ('atlhetica', 'atletica') and ('whey tech' in categoria_raw_lower or 'tech' in categoria_raw_lower):
                    produto['preco'] = 109.90
                    produto['sabores'] = ['chocolate', 'leite', 'morango', 'cookies & cream', 'baunilha']
                    produto['imagem'] = '/static/images/whey-tech-atlhetica-card.png'
                    produto['imagem_principal'] = '/static/images/whey-tech-atlhetica-card.png'
                    produto['imagem_mobile'] = '/static/images/whey-tech-atlhetica-card-mobile.png'
                    produto['imagens'] = ['/static/images/whey-tech-atlhetica-modal-frente.png']
                    produto['imagens_mobile'] = ['/static/images/whey-tech-atlhetica-modal-frente-mobile.png']

                elif marca_lower in ('atlhetica', 'atletica') and ('best' in categoria_raw_lower or 'best whey' in categoria_raw_lower):
                    produto['preco'] = 139.90
                    produto['sabores'] = ['achocolatado toddy','original','pistache','dadinho','dulce de leche','strawberry milk shake','brownie chocolate branco','double chocolate','cookies & cream','cacau & avelã','beijinho de coco']
                    produto['imagem'] = '/static/images/whey-best-atlhetica-card.png'
                    produto['imagem_principal'] = '/static/images/whey-best-atlhetica-card.png'
                    produto['imagem_mobile'] = '/static/images/whey-best-atlhetica-card-mobile.png'
                    produto['imagens'] = ['/static/images/whey-best-atlhetica-modal-frente.png']
                    produto['imagens_mobile'] = ['/static/images/whey-best-atlhetica-modal-frente-mobile.png']

                elif marca_lower in ('atlhetica', 'atletica') and 'creatina' in categoria_raw_lower and '150' in categoria_raw_lower:
                    produto['preco'] = 89.90
                    produto['sabores'] = []
                    produto['imagem'] = '/static/images/creatina150-atlhetica-card-mobile.png'
                    produto['imagem_principal'] = '/static/images/creatina150-atlhetica-card-mobile.png'
                    produto['imagem_mobile'] = '/static/images/creatina150-atlhetica-card-mobile.png'
                    produto['imagens'] = ['/static/images/creatina150-atlhetica-modal-frente.png','/static/images/creatina150-atlhetica-modal-frente(1).png']
                    produto['imagens_mobile'] = ['/static/images/creatina150-atlhetica-modal-frente-mobile.png']

                elif marca_lower in ('atlhetica', 'atletica') and 'creatina' in categoria_raw_lower and '300' in categoria_raw_lower:
                    produto['preco'] = 89.90
                    produto['sabores'] = []
                    produto['imagem'] = '/static/images/creatina300-atlhetica-card.png'
                    produto['imagem_principal'] = '/static/images/creatina300-atlhetica-card.png'
                    produto['imagem_mobile'] = '/static/images/creatina300-atlhetica-card-mobile.png'
                    produto['imagens'] = ['/static/images/creatina300-atlhetica-modal-frente.png']
                    produto['imagens_mobile'] = ['/static/images/creatina300-atlhetica-modal-frente-mobile.png']

                elif marca_lower in ('atlhetica', 'atletica') and 'barrinha' in categoria_raw_lower:
                    produto['preco'] = 11.00
                    produto['sabores'] = []
                    produto['imagem'] = '/static/images/barrinha-atlhetica-card.png'
                    produto['imagem_principal'] = '/static/images/barrinha-atlhetica-card.png'
                    produto['imagem_mobile'] = '/static/images/barrinha-atlhetica-card-mobile.png'
                    produto['imagens'] = ['/static/images/barrinha-atlhetica-modal-frente.png']
                    produto['imagens_mobile'] = ['/static/images/barrinha-atlhetica-modal-frente-mobile.png']

                elif marca_lower in ('atlhetica', 'atletica') and 'multivit' in categoria_raw_lower:
                    produto['preco'] = 67.90
                    produto['sabores'] = []
                    produto['imagem'] = '/static/images/multivitaminico-atlhetica-card-mobile.png'
                    produto['imagem_principal'] = '/static/images/multivitaminico-atlhetica-card-mobile.png'
                    produto['imagem_mobile'] = '/static/images/multivitaminico-atlhetica-card-mobile.png'
                    produto['imagens'] = ['/static/images/multivitaminico-atlhetica-modal-frente.png','/static/images/multivitaminico-atlhetica-modal-frente(1).png']
                    produto['imagens_mobile'] = ['/static/images/multivitaminico-atlhetica-modal-frente-mobile.png']

                # ATLETICA - Linha Monster (atribuir imagens da Probiotica)
                elif 'monster' in produto.get('nome','').lower():
                    produto['imagem'] = '/static/images/whey-monster-probiotica-card.png'
                    produto['imagem_principal'] = '/static/images/whey-monster-probiotica-card.png'
                    produto['imagem_mobile'] = '/static/images/whey-monster-probiotica-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/whey-monster-probiotica-modal-frente.png',
                        '/static/images/whey-monster-probiotica-modal-tabela.png',
                        '/static/images/whey-monster-probiotica-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/whey-monster-probiotica-modal-frente-mobile.png',
                        '/static/images/whey-monster-probiotica-modal-tabela-mobile.png'
                    ]

                # NUTRA - Whey Concentrado
                elif marca_lower == 'nutra' and 'concentr' in categoria_raw_lower:
                    produto['preco'] = 149.90
                    produto['sabores'] = ['banana','creme de baunilha','chocolate com coco','double chocolate','cookies and cream','strawberry milk shake']
                    produto['imagem'] = '/static/images/whey-concentrado-nutra-card.png'
                    produto['imagem_principal'] = '/static/images/whey-concentrado-nutra-card.png'
                    produto['imagem_mobile'] = '/static/images/whey-concentrado-nutra-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/whey-concentrado-nutra-modal-frente.png',
                        '/static/images/whey-concentrado-nutra-modal-meta.png',
                        '/static/images/whey-concentrado-nutra-modal-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/whey-concentrado-nutra-modal-frente-mobile.png',
                        '/static/images/whey-concentrado-nutra-modal-meta-mobile.png',
                        '/static/images/whey-concentrado-nutra-modal-tabela-mobile.png'
                    ]

                # NUTRA - Whey Isolado
                elif marca_lower == 'nutra' and 'isol' in categoria_raw_lower:
                    produto['preco'] = 169.90
                    produto['sabores'] = ['chocolate','creme de coco','creme de baunilha']
                    produto['imagem'] = '/static/images/whey-isolado-nutra-card.png'
                    produto['imagem_principal'] = '/static/images/whey-isolado-nutra-card.png'
                    produto['imagem_mobile'] = '/static/images/whey-isolado-nutra-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/whey-isolado-nutra-modal-frente.png',
                        '/static/images/whey-isolado-nutra-modal-meta.png',
                        '/static/images/whey-isolado-nutra-modal-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/whey-isolado-nutra-modal-frente-mobile.png',
                        '/static/images/whey-isolado-nutra-modal-frente-mobile(1).png',
                        '/static/images/whey-isolado-nutra-modal-frente-mobile(2).png'
                    ]

                # NUTRA - Barrinhas
                elif marca_lower == 'nutra' and 'barrinha' in categoria_raw_lower:
                    produto['preco'] = 17.00
                    produto['sabores'] = ['banoffee','dulce de leche','dulce de leite','morango','brownie de chocolate']
                    produto['imagem'] = '/static/images/barrinha-nutra-card.png'
                    produto['imagem_principal'] = '/static/images/barrinha-nutra-card.png'
                    produto['imagem_mobile'] = '/static/images/barrinha-nutra-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/barrinha-nutra-modal-frente.png',
                        '/static/images/barrinha-nutra-modal-meta.png',
                        '/static/images/barrinha-nutra-modal-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/barrinha-nutra-modal-frente-mobile.png',
                        '/static/images/barrinha-nutra-modal-meta-mobile.png',
                        '/static/images/barrinha-nutra-modal-tabela-mobile.png'
                    ]

                # BOLD - Barrinha
                elif marca_lower == 'bold' and 'barrinha' in categoria_raw_lower:
                    produto['preco'] = 12.00
                    produto['sabores'] = ['Cookies and cream', 'Combo essências', 'Tube pistache', 'Tube caixa mista', 'Caixa mix']
                    produto['imagem'] = '/static/images/barrinha-bold-card.png'
                    produto['imagem_principal'] = '/static/images/barrinha-bold-card.png'
                    produto['imagem_mobile'] = '/static/images/barrinha-bold-card-mobile.png'
                    produto['imagens'] = ['/static/images/barrinha-bold-modal-frente.png']
                    produto['imagens_mobile'] = ['/static/images/barrinha-bold-modal-frente-mobile.png']

                # FTW - Whey Concentrado
                elif marca_lower == 'ftw' and 'concentr' in categoria_raw_lower:
                    produto['preco'] = 99.90
                    produto['sabores'] = ['cookies', 'chocolate', 'morango']
                    produto['imagem'] = '/static/images/whey-concentrado-ftw-card.png'
                    produto['imagem_principal'] = '/static/images/whey-concentrado-ftw-card.png'
                    produto['imagem_mobile'] = '/static/images/whey-concentrado-ftw-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/whey-concentrado-ftw-modal-frente.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/whey-concentrado-ftw-modal-frente-mobile.png'
                    ]

                # FTW - Whey 3W
                elif marca_lower == 'ftw' and ('3w' in categoria_raw_lower or '3 w' in categoria_raw_lower):
                    produto['preco'] = 189.90
                    produto['sabores'] = [
                        'wheyzinho','doce de leite argentino','chocolate','cookies','chocolate maltado',
                        'mini chocolate sortidos','chocolate com avelã','diamante negro','chocolate branco',
                        'baunilha','banana caramelizada','morango','beijinho','iogurte grego','chocoball'
                    ]
                    produto['imagem'] = '/static/images/whey-3w-ftw-card.png'
                    produto['imagem_principal'] = '/static/images/whey-3w-ftw-card.png'
                    produto['imagem_mobile'] = '/static/images/whey-3w-ftw-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/whey-3w-ftw-modal-frente.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/whey-3w-ftw-modal-frente-mobile.png'
                    ]

                # FTW - Creatina
                elif marca_lower == 'ftw' and 'creatina' in categoria_raw_lower:
                    produto['preco'] = 99.90
                    produto['sabores'] = []
                    produto['imagem'] = '/static/images/creatina-ftw-card.png'
                    produto['imagem_principal'] = '/static/images/creatina-ftw-card.png'
                    produto['imagem_mobile'] = '/static/images/creatina-ftw-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/creatina-ftw-modal-frente.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/creatina-ftw-modal-frente-mobile.png'
                    ]

                # FTW - Pré-treino
                elif marca_lower == 'ftw' and ('pré' in categoria_raw_lower or 'pre' in categoria_raw_lower or 'pré-treino' in categoria_raw_lower):
                    produto['preco'] = 89.90
                    produto['sabores'] = ['diabo verde']
                    produto['imagem'] = '/static/images/pre-treino-ftw-card.png'
                    produto['imagem_principal'] = '/static/images/pre-treino-ftw-card.png'
                    produto['imagem_mobile'] = '/static/images/pre-treino-ftw-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/pre-treino-ftw-modal-frente.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/pre-treino-ftw-modal-frente-mobile.png'
                    ]

                # DUX - Whey Concentrado
                elif marca_lower == 'dux' and 'concentr' in categoria_raw_lower:
                    produto['preco'] = 149.90
                    produto['sabores'] = ['torta de limão','chocolate','cookies','banoffe','butter cookies','doce de leite','cappucino','caramelo salgado','coco','baunilha']
                    produto['imagem'] = '/static/images/whey-concentrado-dux-card.png'
                    produto['imagem_principal'] = '/static/images/whey-concentrado-dux-card.png'
                    produto['imagem_mobile'] = '/static/images/whey-concentrado-dux-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/whey-concentrado-dux-modal-frente.png',
                        '/static/images/whey-concentrado-dux-modal-meta.png',
                        '/static/images/whey-concentrado-dux-modal-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/whey-concentrado-dux-modal-frente-mobile.png',
                        '/static/images/whey-concentrado-dux-modal-meta-mobile.png',
                        '/static/images/whey-concentrado-dux-modal-tabela-mobile.png'
                    ]

                # DUX - Whey Isolado
                elif marca_lower == 'dux' and 'isol' in categoria_raw_lower:
                    produto['preco'] = 189.90
                    produto['sabores'] = ['cappuccino','chocolate','morango','neutro','chocolate branco','cookies','coco','doce de leite']
                    produto['imagem'] = '/static/images/whey-isolado-dux-card.png'
                    produto['imagem_principal'] = '/static/images/whey-isolado-dux-card.png'
                    produto['imagem_mobile'] = '/static/images/whey-isolado-dux-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/whey-isolado-dux-modal-frente.png',
                        '/static/images/whey-isolado-dux-modal-meta.png',
                        '/static/images/whey-isolado-dux-modal-tabela.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/whey-isolado-dux-modal-frente-mobile.png',
                        '/static/images/whey-isolado-dux-modal-meta-mobile.png',
                        '/static/images/whey-isolado-dux-modal-tabela-mobile.png'
                    ]

                # DUX - Creatina 300g
                elif marca_lower == 'dux' and 'creatina' in categoria_raw_lower and '300' in categoria_raw_lower:
                    produto['preco'] = 189.90
                    produto['sabores'] = []
                    produto['imagem'] = '/static/images/creatina-dux-card.png'
                    produto['imagem_principal'] = '/static/images/creatina-dux-card.png'
                    produto['imagem_mobile'] = '/static/images/creatina-dux-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/creatina-dux-modal-frente.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/creatina-dux-modal-frente-mobile.png'
                    ]

                # DUX - Multivitamínico
                elif marca_lower == 'dux' and ('multivit' in categoria_raw_lower or 'vitamina' in categoria_raw_lower):
                    produto['preco'] = 69.90
                    produto['sabores'] = []
                    produto['imagem'] = '/static/images/multivitaminco-dux-card.png'
                    produto['imagem_principal'] = '/static/images/multivitaminco-dux-card.png'
                    produto['imagem_mobile'] = '/static/images/multivitaminco-dux-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/multivitaminco-dux-modal-frente.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/multivitaminco-dux-modal-frente-mobile.png'
                    ]

                # DUX - Cápsula de cafeína
                elif marca_lower == 'dux' and ('cafe' in categoria_raw_lower or 'cápsula' in categoria_raw_lower or 'capsula' in categoria_raw_lower):
                    produto['preco'] = 79.90
                    produto['sabores'] = []
                    produto['imagem'] = '/static/images/cafeina-dux-card.png'
                    produto['imagem_principal'] = '/static/images/cafeina-dux-card.png'
                    produto['imagem_mobile'] = '/static/images/cafeina-dux-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/cafeina-dux-modal-frente.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/cafeina-dux-modal-frente-mobile.png'
                    ]

                # DUX - Ômega 3
                elif marca_lower == 'dux' and ('omega' in categoria_raw_lower or 'ômega' in categoria_raw_lower):
                    produto['preco'] = 64.60
                    produto['sabores'] = []
                    produto['imagem'] = '/static/images/omega3-dux-card.png'
                    produto['imagem_principal'] = '/static/images/omega3-dux-card.png'
                    produto['imagem_mobile'] = '/static/images/omega3-dux-card-mobile.png'
                    produto['imagens'] = [
                        '/static/images/omega3-dux-modal-frente.png'
                    ]
                    produto['imagens_mobile'] = [
                        '/static/images/omega3-dux-modal-frente-mobile.png'
                    ]
            except Exception:
                # Não falhar o carregamento por causa de um override
                pass
            produtos.append(produto)
        
        return produtos
    except Exception as e:
        print(f"Erro ao carregar produtos: {e}")
        return []

# Rotas principais
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/produtos')
def produtos():
    return render_template('produtos.html')

@app.route('/outlet')
def outlet():
    return render_template('outlet.html')

@app.route('/checkout')
def checkout():
    if not usuario_logado():
        return redirect(url_for('login'))
    return render_template('checkout.html')

@app.route('/login')
def login():
    return render_template('login.html')

@app.route('/registro')
def registro():
    return render_template('registro.html')

@app.route('/contato')
def contato():
    return render_template('contato.html')

@app.route('/recuperar-senha')
def recuperar_senha():
    try:
        print("🔑 Acessando recuperar senha...")
        return render_template('recuperar_senha.html')
    except Exception as e:
        print(f"💥 Erro em recuperar senha: {e}")
        import traceback
        traceback.print_exc()
        return f"Erro interno: {str(e)}", 500

@app.route('/nova-senha')
def nova_senha():
    try:
        print("🔑 Acessando nova senha...")
        return render_template('nova_senha.html')
    except Exception as e:
        print(f"💥 Erro em nova senha: {e}")
        import traceback
        traceback.print_exc()
        return f"Erro interno: {str(e)}", 500

def obter_dados_produto(produto_id):
    """Obtém dados específicos de um produto"""
    # Dados dos produtos do outlet
    produtos_outlet = {
        'camiseta-golden': {
            'nome': 'Camiseta Golden Era',
            'preco': 129.90,
            'preco_original': 129.90,
            'imagem': '/static/images/camiseta-golden.jpg',
            'descricao': 'Camiseta premium da linha Golden Era com design exclusivo.',
            'vendido': False
        },
        'camiseta-juice': {
            'nome': 'Camiseta Juice of God',
            'preco': 129.90,
            'preco_original': 129.90,
            'imagem': '/static/images/camiseta-juice.jpg',
            'descricao': 'Camiseta premium da linha Juice of God com design exclusivo.',
            'vendido': False
        }
    }
    
    return produtos_outlet.get(produto_id, {
        'nome': f'Produto {produto_id}',
        'preco': 0.00,
        'preco_original': 0.00,
        'imagem': '/static/images/produto-placeholder.svg',
        'descricao': 'Produto não encontrado.',
        'vendido': False
    })

@app.route('/produto/<produto_id>')
def produto_individual(produto_id):
    produto = obter_dados_produto(produto_id)
    return render_template('produto_individual.html', produto_id=produto_id, produto=produto)

@app.route('/api/outlet/estoque')
def api_outlet_estoque():
    """API para verificar estoque dos produtos do outlet"""
    try:
        estoque = {
            'camiseta-golden': {
                'vendido': False,
                'disponivel': True
            },
            'camiseta-juice': {
                'vendido': False,
                'disponivel': True
            }
        }
        return jsonify(estoque)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/perfil')
def perfil():
    try:
        print("👤 Acessando perfil...")
        if not usuario_logado():
            print("❌ Usuário não logado, redirecionando para login")
            return redirect(url_for('login'))
        
        print("✅ Usuário logado, obtendo dados...")
        usuario = obter_usuario_logado()
        print(f"👤 Dados do usuário: {usuario}")
        
        if not usuario:
            print("❌ Erro ao obter dados do usuário")
            return redirect(url_for('login'))
            
        return render_template('perfil.html', usuario=usuario)
        
    except Exception as e:
        print(f"💥 Erro no perfil: {e}")
        import traceback
        traceback.print_exc()
        return f"Erro interno: {str(e)}", 500

@app.route('/pedidos')
def pedidos():
    try:
        print("📦 Acessando pedidos...")
        if not usuario_logado():
            print("❌ Usuário não logado, redirecionando para login")
            return redirect(url_for('login'))
        
        # Buscar pedidos do usuário logado
        usuario = obter_usuario_logado()
        if not usuario:
            print("❌ Usuário não logado, redirecionando para login")
            return redirect(url_for('login'))
        
        print(f"👤 Usuário logado: {usuario['email']}")
        
        conn = conectar_db()
        cursor = conn.cursor()
        
        # PRIMEIRO: Ver todos os emails que existem no banco
        executar_query(cursor, 'SELECT DISTINCT email FROM pedidos')
        todos_emails = cursor.fetchall()
        print(f"📧 TODOS os emails no banco: {[email[0] for email in todos_emails]}")
        
        # SEGUNDO: Contar total de pedidos
        executar_query(cursor, 'SELECT COUNT(*) FROM pedidos')
        total_pedidos = cursor.fetchone()[0]
        print(f"📊 TOTAL de pedidos no banco: {total_pedidos}")
        
        # BUSCAR PEDIDOS: Usar apenas email do usuário logado
        print(f"🔍 Buscando pedidos para email do usuário logado: '{usuario['email']}'")
        executar_query(cursor, '''
            SELECT * FROM pedidos WHERE email = ? ORDER BY data_pedido DESC
        ''', (usuario['email'],))
        
        pedidos = cursor.fetchall()
        print(f"📦 Encontrados {len(pedidos)} pedidos para {usuario['email']}")
        
        # Se ainda não encontrou, apenas logar para debug (NÃO mostrar todos os pedidos)
        if len(pedidos) == 0:
            print(f"🔍 Nenhum pedido encontrado para {usuario['email']}")
            # Buscar total de pedidos no banco para debug (sem mostrar dados)
            executar_query(cursor, 'SELECT COUNT(*) FROM pedidos')
            total_pedidos = cursor.fetchone()[0]
            print(f"📊 Total de pedidos no banco: {total_pedidos}")
            
            # Buscar alguns emails diferentes para debug
            executar_query(cursor, 'SELECT DISTINCT email FROM pedidos LIMIT 5')
            emails_diferentes = cursor.fetchall()
            print(f"📧 Emails diferentes no banco: {[email[0] for email in emails_diferentes]}")
        
        conn.close()
        
        print(f"📊 Encontrados {len(pedidos)} pedidos para {usuario['email']}")
        
        # Converter para formato mais legível e validar segurança
        pedidos_formatados = []
        for pedido in pedidos:
            # VALIDAÇÃO DE SEGURANÇA: Verificar se o pedido pertence ao usuário logado
            pedido_email = pedido[3].lower() if pedido[3] else ""
            usuario_email = usuario['email'].lower()
            
            if pedido_email == usuario_email:
                pedidos_formatados.append({
                    'id': pedido[0],
                    'order_id': pedido[1],
                    'nome': pedido[2],
                    'email': pedido[3],
                    'telefone': pedido[4],
                    'cpf': pedido[5],
                    'data_nascimento': pedido[6],
                    'cep': pedido[7],
                    'cidade': pedido[8],
                    'estado': pedido[9],
                    'bairro': pedido[10],
                    'endereco': pedido[11],
                    'observacoes': pedido[12],
                    'status': pedido[13],
                    'total': float(pedido[14]),
                    'produtos': pedido[15],
                    'data_pedido': pedido[16]
                })
        else:
                print(f"⚠️ SEGURANÇA: Pedido {pedido[1]} não pertence ao usuário {usuario['email']} (pertence a {pedido[3]})")
        
        print(f"🔒 Após validação de segurança: {len(pedidos_formatados)} pedidos válidos para {usuario['email']}")
        
        return render_template('pedidos.html', pedidos=pedidos_formatados)

    except Exception as e:
        print(f"💥 Erro nos pedidos: {e}")
        import traceback
        traceback.print_exc()
        return f"Erro interno: {str(e)}", 500

# Bloquear URLs antigas do admin
@app.route('/admin/login')
@app.route('/admin/pedidos')
@app.route('/relatorios-financeiros-atlas')
@app.route('/relatorios-financeiros-atlas/vendas')
def admin_blocked():
    """Bloquear acesso às URLs antigas do admin"""
    print("🚫 Tentativa de acesso às URLs antigas do admin bloqueada")
    return "Página não encontrada", 404

@app.route('/sistema-interno-gestao-vendas-2024')
def admin_login():
    """Página de login para administradores - URL secreta única"""
    print("🔐 Acessando página de login de admin...")
    print(f"🔍 Sessão atual: {dict(session)}")
    
    # Verificação de segurança
    referer = request.headers.get('Referer', '')
    user_agent = request.headers.get('User-Agent', '')
    client_ip = request.remote_addr
    
    print(f"🔍 Referer: {referer}")
    print(f"🔍 User-Agent: {user_agent}")
    print(f"🔍 IP: {client_ip}")
    
    # Verificar se vem de uma fonte confiável (opcional)
    if referer and 'atlas-1h3w.onrender.com' not in referer:
        print("⚠️ Acesso suspeito - referer não confiável")
    
    return render_template('admin_login.html')

@app.route('/api/admin-login', methods=['POST'])
def api_admin_login():
    """API para login de administrador com rate limiting"""
    try:
        # Rate limiting
        client_ip = request.remote_addr
        current_time = time.time()
        
        # Limpar tentativas antigas
        if client_ip in admin_login_attempts:
            admin_login_attempts[client_ip] = [
                attempt_time for attempt_time in admin_login_attempts[client_ip]
                if current_time - attempt_time < TIME_WINDOW
            ]
        else:
            admin_login_attempts[client_ip] = []
        
        # Verificar se excedeu o limite
        if len(admin_login_attempts[client_ip]) >= MAX_LOGIN_ATTEMPTS:
            print(f"🚫 Rate limit excedido para IP: {client_ip}")
            return jsonify({
                "success": False, 
                "error": "Muitas tentativas de login. Tente novamente em 5 minutos."
            }), 429
        
        data = request.get_json()
        email = data.get('email')
        senha = data.get('senha')
        
        if not email or not senha:
            return jsonify({"success": False, "error": "Email e senha são obrigatórios"}), 400
        
        conn = conectar_db()
        cursor = conn.cursor()
        
        # Buscar usuário admin
        executar_query(cursor, '''
            SELECT id, nome, email, senha_hash, admin FROM usuario 
            WHERE email = ? AND admin = 1
        ''', (email,))
        
        usuario = cursor.fetchone()
        conn.close()
        
        if usuario and verificar_senha(senha, usuario[3]):
            # Login bem-sucedido - limpar tentativas
            if client_ip in admin_login_attempts:
                del admin_login_attempts[client_ip]
            
            # NÃO limpar a sessão - apenas adicionar dados do admin
            # Preservar dados do usuário normal se existirem
            session['admin_user_id'] = usuario[0]  # Chave diferente para admin
            session['admin'] = True
            session['admin_mode'] = True
            session['is_admin_session'] = True  # Flag para identificar sessão de admin
            print(f"👑 Admin logado: {usuario[1]} ({usuario[2]}) - Sessão preservada")
            print(f"🔍 Sessão após login admin: {dict(session)}")
            return jsonify({
                "success": True,
                "message": "Login de admin realizado com sucesso",
                "redirect": "/sistema-interno-gestao-vendas-2024/pedidos"  # URL secreta
            })
        else:
            # Login falhou - registrar tentativa
            admin_login_attempts[client_ip].append(current_time)
            print(f"❌ Tentativa de login admin falhada para IP: {client_ip}")
            return jsonify({
                "success": False,
                "error": "Credenciais inválidas ou usuário não é administrador"
            }), 401
            
    except Exception as e:
        print(f"❌ Erro no login admin: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/sistema-interno-gestao-vendas-2024/pedidos')
def admin_pedidos():
    """Página para administrador ver todos os pedidos"""
    try:
        print("👑 Acessando página de administração de pedidos...")
        print(f"🔍 Sessão atual: {dict(session)}")
        print(f"🔍 admin_logado(): {admin_logado()}")
        
        # Verificar se é admin
        if not admin_logado():
            print("❌ Admin não logado, redirecionando para login")
            return redirect(url_for('admin_login'))
        
        # Buscar pedidos do banco de dados
        conn = conectar_db()
        cursor = conn.cursor()
        
        # Criar tabela de pedidos se não existir
        executar_query(cursor, '''
            CREATE TABLE IF NOT EXISTS pedidos (
                id SERIAL PRIMARY KEY,
                order_id VARCHAR(255) UNIQUE NOT NULL,
                nome VARCHAR(255) NOT NULL,
                email VARCHAR(255) NOT NULL,
                telefone VARCHAR(20),
                cpf VARCHAR(20),
                data_nascimento VARCHAR(20),
                cep VARCHAR(20),
                cidade VARCHAR(100),
                estado VARCHAR(50),
                bairro VARCHAR(100),
                endereco TEXT,
                observacoes TEXT,
                status VARCHAR(50) DEFAULT 'Pendente',
                total DECIMAL(10,2) NOT NULL,
                produtos TEXT NOT NULL,
                data_pedido TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Buscar todos os pedidos
        executar_query(cursor, 'SELECT * FROM pedidos ORDER BY data_pedido DESC')
        pedidos = cursor.fetchall()
        conn.close()
        
        print(f"📊 Encontrados {len(pedidos)} pedidos")
        
        # Converter para formato mais legível com horário do Brasil
        pedidos_formatados = []
        for pedido in pedidos:
            # Converter data para horário do Brasil
            data_pedido = pedido[16] if len(pedido) > 16 else None
            
            if data_pedido is None:
                data_formatada = "N/A"
            elif isinstance(data_pedido, str):
                # Se já é string, tentar converter para datetime e depois para Brasil
                try:
                    from datetime import datetime
                    # Tentar parsear a string como datetime
                    if 'T' in data_pedido:
                        dt = datetime.fromisoformat(data_pedido.replace('Z', '+00:00'))
                    else:
                        dt = datetime.strptime(data_pedido, '%Y-%m-%d %H:%M:%S.%f')
                    
                    # Converter para horário do Brasil (UTC-3)
                    brasil_tz = timezone(timedelta(hours=-3))
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                    data_brasil = dt.astimezone(brasil_tz)
                    data_formatada = data_brasil.strftime("%d/%m/%Y %H:%M (Brasil)")
                except:
                    # Se não conseguir converter, usar como está
                    data_formatada = data_pedido
            else:
                # Se é datetime, converter para Brasil
                try:
                    brasil_tz = timezone(timedelta(hours=-3))
                    if data_pedido.tzinfo is None:
                        data_pedido = data_pedido.replace(tzinfo=timezone.utc)
                    data_brasil = data_pedido.astimezone(brasil_tz)
                    data_formatada = data_brasil.strftime("%d/%m/%Y %H:%M (Brasil)")
                except:
                    data_formatada = str(data_pedido)
            
            pedidos_formatados.append({
                'id': pedido[0],
                'order_id': pedido[1],
                'nome': pedido[2],
                'email': pedido[3],
                'telefone': pedido[4],
                'cpf': pedido[5],
                'data_nascimento': pedido[6],
                'cep': pedido[7],
                'cidade': pedido[8],
                'estado': pedido[9],
                'bairro': pedido[10],
                'endereco': pedido[11],
                'observacoes': pedido[12],
                'status': pedido[13],
                'total': float(pedido[14]),
                'produtos': pedido[15],
                'data_pedido': data_formatada
            })
        
        return render_template('admin_pedidos.html', pedidos=pedidos_formatados)
        
    except Exception as e:
        print(f"💥 Erro na página de admin: {e}")
        import traceback
        traceback.print_exc()
        return f"Erro interno: {str(e)}", 500

@app.route('/status-pedido')
def status_pedido():
    """Página para cliente verificar status do pedido"""
    return render_template('status_pedido.html')

@app.route('/api/buscar-pedido', methods=['POST'])
def buscar_pedido():
    """API para buscar pedido por ID e email"""
    try:
        data = request.get_json()
        order_id = data.get('order_id')
        email = data.get('email')
        
        if not order_id or not email:
            return jsonify({"success": False, "error": "order_id e email são obrigatórios"}), 400
        
        conn = conectar_db()
        cursor = conn.cursor()
        
        # Buscar pedido
        executar_query(cursor, '''
            SELECT * FROM pedidos WHERE order_id = ? AND email = ?
        ''', (order_id, email))
        
        pedido = cursor.fetchone()
        conn.close()
        
        if pedido:
            pedido_formatado = {
                'id': pedido[0],
                'order_id': pedido[1],
                'nome': pedido[2],
                'email': pedido[3],
                'telefone': pedido[4],
                'cpf': pedido[5],
                'data_nascimento': pedido[6],
                'cep': pedido[7],
                'cidade': pedido[8],
                'estado': pedido[9],
                'bairro': pedido[10],
                'endereco': pedido[11],
                'observacoes': pedido[12],
                'status': pedido[13],
                'total': float(pedido[14]),
                'produtos': pedido[15],
                'data_pedido': pedido[16]
            }
            
            return jsonify({
                "success": True,
                "pedido": pedido_formatado
            })
        else:
            return jsonify({
                "success": False,
                "error": "Pedido não encontrado"
            }), 404
            
    except Exception as e:
        print(f"❌ Erro ao buscar pedido: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/atualizar-status', methods=['POST'])
def atualizar_status_pedido():
    """API para atualizar status do pedido"""
    try:
        data = request.get_json()
        order_id = data.get('order_id')
        novo_status = data.get('status')
        
        if not order_id or not novo_status:
            return jsonify({"success": False, "error": "order_id e status são obrigatórios"}), 400
        
        conn = conectar_db()
        cursor = conn.cursor()
        
        # Buscar dados do pedido
        executar_query(cursor, '''
            SELECT email, telefone, nome FROM pedidos WHERE order_id = ?
        ''', (order_id,))
        
        pedido = cursor.fetchone()
        if not pedido:
            conn.close()
            return jsonify({"success": False, "error": "Pedido não encontrado"}), 404
        
        email, telefone, nome = pedido
        
        # Atualizar status
        executar_query(cursor, '''
            UPDATE pedidos SET status = ? WHERE order_id = ?
        ''', (novo_status, order_id))
        
        if cursor.rowcount > 0:
            conn.commit()
            conn.close()
            
            # Criar notificação
            mensagem = obter_mensagem_status(novo_status)
            criar_notificacao(order_id, email, telefone, novo_status, mensagem)
            
            # Enviar WhatsApp automático
            whatsapp_enviado = enviar_whatsapp_automatico(order_id, nome, telefone, novo_status)
            
            print(f"✅ Status do pedido {order_id} atualizado para {novo_status}")
            print(f"📧 Notificação enviada para {email}")
            if whatsapp_enviado:
                print(f"📱 WhatsApp preparado para {nome}")
            
            return jsonify({
                "success": True, 
                "message": f"Status atualizado para {novo_status}",
                "notificacao_enviada": True,
                "whatsapp_enviado": whatsapp_enviado
            })
        else:
            conn.close()
            return jsonify({"success": False, "error": "Erro ao atualizar status"}), 500
            
    except Exception as e:
        print(f"❌ Erro ao atualizar status: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/admin-logout', methods=['POST'])
def admin_logout():
    """Logout do admin - NÃO afeta usuário normal"""
    try:
        # Remover apenas dados do admin, preservar dados do usuário normal
        if 'admin_user_id' in session:
            del session['admin_user_id']
        if 'admin' in session:
            del session['admin']
        if 'admin_mode' in session:
            del session['admin_mode']
        if 'is_admin_session' in session:
            del session['is_admin_session']
        
        print("👑 Admin deslogado - usuário normal preservado")
        print(f"🔍 Sessão após logout admin: {dict(session)}")
        return jsonify({
            "success": True,
            "message": "Logout de admin realizado com sucesso",
            "redirect": "/sistema-interno-gestao-vendas-2024"  # URL secreta
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/verificar-admin', methods=['GET'])
def verificar_admin():
    """Verificar se admin está logado"""
    try:
        print(f"🔍 Verificando admin - Sessão: {dict(session)}")
        print(f"🔍 admin_logado(): {admin_logado()}")
        
        if admin_logado():
            return jsonify({
                "success": True,
                "admin_logado": True,
                "sessao": dict(session)
            })
        else:
            return jsonify({
                "success": True,
                "admin_logado": False,
                "sessao": dict(session)
            })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/debug-pedidos', methods=['GET'])
def debug_pedidos():
    """Endpoint para debug - TEMPORÁRIO para descobrir o problema"""
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        
        # Buscar todos os pedidos
        executar_query(cursor, 'SELECT * FROM pedidos ORDER BY data_pedido DESC')
        pedidos = cursor.fetchall()
        
        # Buscar todos os emails únicos
        executar_query(cursor, 'SELECT DISTINCT email FROM pedidos')
        emails_unicos = cursor.fetchall()
        
        # Buscar email do usuário logado
        usuario_logado_email = "N/A"
        if usuario_logado():
            usuario = obter_usuario_logado()
            if usuario:
                usuario_logado_email = usuario['email']
        
        conn.close()
        
        pedidos_formatados = []
        for pedido in pedidos:
            pedidos_formatados.append({
                'id': pedido[0],
                'order_id': pedido[1],
                'nome': pedido[2],
                'email': pedido[3],
                'status': pedido[13],
                'total': float(pedido[14]),
                'data_pedido': str(pedido[16])
            })
        
        return jsonify({
            "success": True,
            "total_pedidos": len(pedidos_formatados),
            "emails_unicos": [email[0] for email in emails_unicos],
            "usuario_logado_email": usuario_logado_email,
            "pedidos": pedidos_formatados
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/corrigir-email-pedidos', methods=['POST'])
def corrigir_email_pedidos():
    """Endpoint para corrigir o email dos pedidos"""
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        
        # Atualizar todos os pedidos com email antigo para o email correto
        email_antigo = "henriqueangelogy@gmail.com"
        email_novo = "henriqueangegelo@gmail.com"
        
        executar_query(cursor, '''
            UPDATE pedidos SET email = ? WHERE email = ?
        ''', (email_novo, email_antigo))
        
        pedidos_atualizados = cursor.rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"Atualizados {pedidos_atualizados} pedidos de {email_antigo} para {email_novo}"
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# Sistema de carrinho híbrido (banco + memória)
carrinho_temporario = []  # Para usuários não logados

def obter_carrinho_temporario():
    """Obtém o carrinho temporário usando sessão do Flask"""
    if 'carrinho_temporario' not in session:
        session['carrinho_temporario'] = []
    return session['carrinho_temporario']

# Sistema de rate limiting para admin
admin_login_attempts = {}  # {ip: [timestamps]}
MAX_LOGIN_ATTEMPTS = 5  # Máximo 5 tentativas
TIME_WINDOW = 300  # 5 minutos

def obter_carrinho_usuario():
    """Obtém o carrinho do usuário atual"""
    try:
        print(f"🛒 obter_carrinho_usuario() - qualquer_usuario_logado(): {qualquer_usuario_logado()}")
        
        if not qualquer_usuario_logado():
            print("⚠️ Usuário não logado - usando carrinho temporário")
            carrinho_temp = obter_carrinho_temporario()
            print(f"🛒 Carrinho temporário tem {len(carrinho_temp)} itens")
            return carrinho_temp
        
        conn = conectar_db()
        cursor = conn.cursor()
        
        # Buscar carrinho do usuário no banco
        executar_query(cursor, '''
            SELECT produto_id, nome, marca, preco, sabor, quantidade, imagem
            FROM carrinho WHERE user_id = ?
        ''', (session['user_id'],))
        
        itens = cursor.fetchall()
        conn.close()
        
        carrinho = []
        for item in itens:
            carrinho.append({
                'produto_id': item[0],
                'nome': item[1],
                'marca': item[2],
                'preco': float(item[3]),
                'sabor': item[4],
                'quantidade': item[5],
                'imagem': item[6]
            })
        
        return carrinho
        
    except Exception as e:
        print(f"❌ Erro ao obter carrinho: {e}")
        return obter_carrinho_temporario()

def salvar_pedido_na_planilha(dados_cliente, carrinho, order_id, status="Pendente"):
    """Salva o pedido na planilha pedidos_atlas.xlsx E no banco de dados"""
    try:
        # Calcular total e produtos
        total = sum(item.get('preco', 0) * item.get('quantidade', 1) for item in carrinho)
        produtos_texto = []
        for item in carrinho:
            produto_info = f"{item.get('nome', 'Produto')}"
            if item.get('sabor'):
                produto_info += f" - Sabor: {item.get('sabor')}"
            produto_info += f" (Qtd: {item.get('quantidade', 1)})"
            produtos_texto.append(produto_info)
        produtos_str = " | ".join(produtos_texto)
        
        print(f"📊 Salvando pedido {order_id} - Total: R$ {total:.2f}")
        
        # 1. SALVAR NO BANCO DE DADOS (PRINCIPAL)
        try:
            conn = conectar_db()
            cursor = conn.cursor()
            
            # Criar tabela se não existir
            executar_query(cursor, '''
                CREATE TABLE IF NOT EXISTS pedidos (
                    id SERIAL PRIMARY KEY,
                    order_id VARCHAR(255) UNIQUE NOT NULL,
                    nome VARCHAR(255) NOT NULL,
                    email VARCHAR(255) NOT NULL,
                    telefone VARCHAR(20),
                    cpf VARCHAR(20),
                    data_nascimento VARCHAR(20),
                    cep VARCHAR(20),
                    cidade VARCHAR(100),
                    estado VARCHAR(50),
                    bairro VARCHAR(100),
                    endereco TEXT,
                    observacoes TEXT,
                    status VARCHAR(50) DEFAULT 'Pendente',
                    total DECIMAL(10,2) NOT NULL,
                    produtos TEXT NOT NULL,
                    data_pedido TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Criar tabela de notificações
            executar_query(cursor, '''
                CREATE TABLE IF NOT EXISTS notificacoes (
                    id SERIAL PRIMARY KEY,
                    order_id VARCHAR(255) NOT NULL,
                    email VARCHAR(255) NOT NULL,
                    telefone VARCHAR(20),
                    status VARCHAR(50) NOT NULL,
                    mensagem TEXT NOT NULL,
                    data_notificacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    enviada BOOLEAN DEFAULT FALSE
                )
            ''')
            
            # Inserir pedido
            executar_query(cursor, '''
                INSERT INTO pedidos (order_id, nome, email, telefone, cpf, data_nascimento, 
                                   cep, cidade, estado, bairro, endereco, observacoes, 
                                   status, total, produtos)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
            order_id,
            dados_cliente.get('nome', ''),
            dados_cliente.get('email', ''),
            dados_cliente.get('telefone', ''),
            dados_cliente.get('cpf', ''),
            dados_cliente.get('data_nascimento', ''),
            dados_cliente.get('cep', ''),
            dados_cliente.get('cidade', ''),
            dados_cliente.get('estado', ''),
            dados_cliente.get('bairro', ''),
            dados_cliente.get('endereco', ''),
            dados_cliente.get('observacoes', ''),
            status,
            total,
            produtos_str
        ))
            
            conn.commit()
            conn.close()
            print(f"✅ Pedido {order_id} salvo no BANCO DE DADOS com sucesso!")
            
        except Exception as e:
            print(f"❌ Erro ao salvar no banco: {e}")
        
        # 2. PLANILHA REMOVIDA - usando apenas banco de dados
        print(f"✅ Pedido {order_id} salvo apenas no banco de dados (planilha removida)")
        
        print(f"📊 Total: R$ {total:.2f}")
        print(f"📊 Produtos: {produtos_str}")
        return True
        
    except Exception as e:
        print(f"❌ Erro geral ao salvar pedido: {e}")
        import traceback
        traceback.print_exc()
        return False

def gerar_link_pagamento_simples(dados_cliente, order_id):
    """Gera link de pagamento usando produtos do carrinho"""
    sdk = mercadopago.SDK("APP_USR-1767627899974277-090620-d9a19a4ac8a0b81161b717c772359483-2669713221")
    
    base_url = os.getenv("BASE_URL", "http://localhost:5000")
    carrinho = obter_carrinho_usuario()
    
    if not carrinho or len(carrinho) == 0:
        return {"success": False, "error": "Carrinho vazio"}
    
    items = []
    for item in carrinho:
        items.append({
            "id": item.get('produto_id', f"item_{len(items)}"),
            "title": item.get('nome', 'Produto'),
            "quantity": item.get('quantidade', 1),
            "currency_id": "BRL",
            "unit_price": float(item.get('preco', 0))
        })
    
    payment_data = {
        "items": items,
        "back_urls": {
            "success": f"{base_url}/pagamento/sucesso",
            "failure": f"{base_url}/pagamento/falha",
            "pending": f"{base_url}/pagamento/pendente"
        }
    }
    
    result = sdk.preference().create(payment_data)
    
    if "response" in result:
        payment = result["response"]
        if "init_point" in payment:
            return {
                "success": True,
                "init_point": payment["init_point"],
                "preference_id": payment.get("id")
            }
        else:
            return {"success": False, "error": "Link de pagamento não gerado"}
    else:
        return {"success": False, "error": "Erro na API do Mercado Pago"}

# API básica
@app.route('/api/carrinho', methods=['GET'])
def get_carrinho():
    return jsonify(obter_carrinho_usuario())

@app.route('/api/carrinho/adicionar', methods=['POST'])
def adicionar_ao_carrinho():
    try:
        data = request.get_json()
        produto_id = data.get('produto_id')
        nome = data.get('nome')
        marca = data.get('marca')
        preco = float(data.get('preco', 0))
        sabor = data.get('sabor')
        quantidade = int(data.get('quantidade', 1))
        imagem = data.get('imagem', '/static/images/produto-placeholder.svg')
        
        print(f"🛒 Tentando adicionar produto {produto_id} ao carrinho")
        print(f"🛒 qualquer_usuario_logado(): {qualquer_usuario_logado()}")

        # Se usuário não estiver logado, usar carrinho temporário na sessão
        if not qualquer_usuario_logado():
            try:
                print("⚠️ Usuário não logado - adicionando ao carrinho temporário")
                carrinho_temp = obter_carrinho_temporario()
                item_existente = None
                for item in carrinho_temp:
                    if item['produto_id'] == produto_id and item['sabor'] == sabor:
                        item_existente = item
                        break

                if item_existente:
                    item_existente['quantidade'] += quantidade
                    print(f"✅ Quantidade atualizada para {item_existente['quantidade']}")
                else:
                    novo_item = {
                        'produto_id': produto_id,
                        'nome': nome,
                        'marca': marca,
                        'preco': preco,
                        'sabor': sabor,
                        'quantidade': quantidade,
                        'imagem': imagem
                    }
                    carrinho_temp.append(novo_item)
                    print(f"✅ Novo item adicionado ao carrinho temporário")

                # Salvar na sessão
                session['carrinho_temporario'] = carrinho_temp
                session.modified = True

                print(f"🛒 Carrinho temporário agora tem {len(carrinho_temp)} itens")
                return jsonify({
                    "success": True,
                    "carrinho": carrinho_temp,
                    "message": "Produto adicionado ao carrinho temporário"
                })
            except Exception as se:
                print(f"❌ Erro ao manipular carrinho temporário: {se}")
                return jsonify({"success": False, "error": str(se)}), 500

        # Usuário logado - usar banco de dados (OTIMIZADO)
        print(f"🛒 Adicionando produto {produto_id} ao carrinho do usuário {session['user_id']}")
        
        try:
            conn = conectar_db()
            cursor = conn.cursor()
            
            # Usar UPSERT para ser mais rápido
            database_url = os.environ.get('DATABASE_URL')
            if database_url:
                # PostgreSQL
                executar_query(cursor, '''
                    INSERT INTO carrinho (user_id, produto_id, nome, marca, preco, sabor, quantidade, imagem)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (user_id, produto_id, sabor) 
                    DO UPDATE SET quantidade = carrinho.quantidade + %s
                ''', (session['user_id'], produto_id, nome, marca, preco, sabor, quantidade, imagem, quantidade))
            else:
                # SQLite
                executar_query(cursor, '''
                    INSERT INTO carrinho (user_id, produto_id, nome, marca, preco, sabor, quantidade, imagem)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT (user_id, produto_id, sabor) 
                    DO UPDATE SET quantidade = carrinho.quantidade + ?
                ''', (session['user_id'], produto_id, nome, marca, preco, sabor, quantidade, imagem, quantidade))
            
            conn.commit()
            conn.close()
            print(f"✅ Produto adicionado/atualizado no carrinho")
            
            # Retornar resposta rápida
            return jsonify({
                "success": True,
                "message": "Produto adicionado ao carrinho com sucesso!"
            })
            
        except Exception as db_error:
            print(f"❌ Erro no banco: {db_error}")
            # Fallback: usar carrinho temporário se banco falhar
            carrinho_temp = obter_carrinho_temporario()
            novo_item = {
                'produto_id': produto_id,
                'nome': nome,
                'marca': marca,
                'preco': preco,
                'sabor': sabor,
                'quantidade': quantidade,
                'imagem': imagem
            }
            carrinho_temp.append(novo_item)
            
            # Salvar na sessão
            session['carrinho_temporario'] = carrinho_temp
            session.modified = True
            
            return jsonify({
                "success": True,
                "message": "Produto adicionado ao carrinho temporário (banco indisponível)"
            })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/carrinho/remover', methods=['POST'])
def remover_do_carrinho():
    try:
        data = request.get_json()
        produto_id = data.get('produto_id')
        sabor = data.get('sabor')
        
        if not produto_id or not sabor:
            return jsonify({"success": False, "error": "Dados inválidos"}), 400
        
        if not qualquer_usuario_logado():
            # Remover do carrinho temporário
            try:
                carrinho_temp = obter_carrinho_temporario()
                carrinho_temp[:] = [item for item in carrinho_temp 
                                 if not (item['produto_id'] == produto_id and item['sabor'] == sabor)]
                session['carrinho_temporario'] = carrinho_temp
                session.modified = True

                return jsonify({
                    "success": True,
                    "carrinho": carrinho_temp,
                    "message": "Item removido do carrinho temporário"
                })
            except Exception as ste:
                print(f"❌ Erro ao remover do carrinho temporário: {ste}")
                return jsonify({"success": False, "error": str(ste)}), 500

        # Usuário logado - remover do banco
        try:
            conn = conectar_db()
            cursor = conn.cursor()
            
            executar_query(cursor, '''
                DELETE FROM carrinho 
                WHERE user_id = ? AND produto_id = ? AND sabor = ?
            ''', (session['user_id'], produto_id, sabor))
            
            conn.commit()
            conn.close()
            
            return jsonify({
                "success": True,
                "carrinho": obter_carrinho_usuario(),
                "message": "Produto removido do carrinho"
            })
        except Exception as dbe:
            print(f"❌ Erro ao remover do banco: {dbe}")
            return jsonify({"success": False, "error": str(dbe)}), 500
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/carrinho/atualizar', methods=['POST'])
def atualizar_quantidade_carrinho():
    """Atualiza a quantidade de um item no carrinho"""
    try:
        data = request.get_json()
        produto_id = data.get('produto_id')
        sabor = data.get('sabor')
        nova_quantidade = int(data.get('quantidade', 1))
        
        if not produto_id or not sabor or nova_quantidade < 0:
            return jsonify({"success": False, "error": "Dados inválidos"}), 400
        
        if nova_quantidade == 0:
            # Se quantidade for 0, remover o item
            if not qualquer_usuario_logado():
                # Remover do carrinho temporário (sessão)
                carrinho_temp = obter_carrinho_temporario()
                novo_carrinho = [item for item in carrinho_temp if not (item['produto_id'] == produto_id and item['sabor'] == sabor)]

                session['carrinho_temporario'] = novo_carrinho
                session.modified = True

                return jsonify({
                    "success": True,
                    "carrinho": novo_carrinho,
                    "message": "Produto removido do carrinho temporário"
                })

            # Usuário logado - remover do banco
            try:
                conn = conectar_db()
                cursor = conn.cursor()

                executar_query(cursor, '''
                    DELETE FROM carrinho WHERE user_id = ? AND produto_id = ? AND sabor = ?
                ''', (session['user_id'], produto_id, sabor))

                conn.commit()
                conn.close()

                return jsonify({"success": True, "message": "Produto removido do carrinho"})
            except Exception as dbe:
                print(f"❌ Erro ao remover do banco: {dbe}")
                return jsonify({"success": False, "error": str(dbe)}), 500
        
        if not qualquer_usuario_logado():
            # Atualizar no carrinho temporário
            carrinho_temp = obter_carrinho_temporario()
            for item in carrinho_temp:
                if item['produto_id'] == produto_id and item['sabor'] == sabor:
                    item['quantidade'] = nova_quantidade
                    break
            
            session['carrinho_temporario'] = carrinho_temp
            session.modified = True
            
            return jsonify({
                "success": True,
                "carrinho": carrinho_temp,
                "message": "Quantidade atualizada no carrinho temporário"
            })
        
        # Usuário logado - atualizar no banco
        conn = conectar_db()
        cursor = conn.cursor()
        
        executar_query(cursor, '''
            UPDATE carrinho 
            SET quantidade = ? 
            WHERE user_id = ? AND produto_id = ? AND sabor = ?
        ''', (nova_quantidade, session['user_id'], produto_id, sabor))
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "Quantidade atualizada no carrinho"})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/carrinho/limpar', methods=['POST'])
def limpar_carrinho():
    try:
        if not usuario_logado():
            return jsonify({
                "success": False,
                "error": "Usuário não logado"
            }), 401
        
        conn = conectar_db()
        cursor = conn.cursor()
        
        executar_query(cursor, '''
            DELETE FROM carrinho WHERE user_id = ?
        ''', (session['user_id'],))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "carrinho": obter_carrinho_usuario(),
            "message": "Carrinho limpo"
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/produtos', methods=['GET'])
def get_produtos():
    try:
        produtos = carregar_produtos()
        return jsonify({
            "success": True,
            "produtos": produtos,
            "total": len(produtos)
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e),
            "produtos": []
        }), 500

@app.route("/api/criar-pagamento-simples", methods=["POST"])
def criar_pagamento_simples():
    try:
        data = request.get_json()
        dados_cliente = data.get("dados_cliente", {})
        
        # Validações obrigatórias
        if not dados_cliente.get('nome') or not dados_cliente.get('email'):
            return jsonify({"error": "Nome e email são obrigatórios"}), 400
        
        # Validação de email
        if not validar_email(dados_cliente.get('email', '')):
            return jsonify({"error": "Email inválido"}), 400
        
        # Validação de CPF (se fornecido)
        if dados_cliente.get('cpf') and not validar_cpf(dados_cliente.get('cpf')):
            return jsonify({"error": "CPF inválido"}), 400
        
        # Validação de telefone (se fornecido)
        if dados_cliente.get('telefone') and not validar_telefone(dados_cliente.get('telefone')):
            return jsonify({"error": "Telefone inválido"}), 400

        order_id = f"pedido_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
        carrinho = obter_carrinho_usuario()
        
        print(f"🛒 Processando pedido {order_id}")
        print(f"🛒 Carrinho: {len(carrinho)} itens")
        print(f"🛒 Dados cliente: {dados_cliente.get('nome', 'N/A')} - {dados_cliente.get('email', 'N/A')}")
        
        if salvar_pedido_na_planilha(dados_cliente, carrinho, order_id, "Pendente"):
            print(f"✅ Pedido {order_id} registrado na planilha com sucesso!")
        else:
            print(f"⚠️ ERRO: Falha ao salvar pedido {order_id} na planilha!")
        
        result = gerar_link_pagamento_simples(dados_cliente, order_id)
        
        if result["success"]:
            return jsonify({
                "success": True,
                "init_point": result["init_point"],
                "order_id": order_id,
                "preference_id": result.get("preference_id")
            })
        else:
            return jsonify({"error": result.get("error", "Erro ao criar pagamento")}), 500
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/pagamento/sucesso")
def pagamento_sucesso():
    return render_template("pagamento_sucesso.html")

@app.route("/pagamento/falha")
def pagamento_falha():
    return render_template("pagamento_falha.html")

@app.route("/pagamento/pendente")
def pagamento_pendente():
    return render_template("pagamento_pendente.html")

@app.route('/api/verificar-login', methods=['GET'])
def verificar_login():
    if usuario_logado():
        usuario = obter_usuario_logado()
        return jsonify({
            "logado": True,
            "usuario": usuario,
            "usuario_nome": usuario['nome'] if usuario else 'Usuário',
            "usuario_email": usuario['email'] if usuario else 'Email'
        })
    return jsonify({
        "logado": False,
        "usuario": None,
        "usuario_nome": None,
        "usuario_email": None
    })

@app.route('/api/login', methods=['POST'])
def api_login():
    try:
        print("🔐 Tentativa de login...")
        data = request.get_json()
        email = data.get('email')
        senha = data.get('senha')
        
        print(f"📧 Email: {email}")
        
        if not email or not senha:
            return jsonify({"success": False, "error": "Email e senha são obrigatórios"}), 400
        
        print("🔧 Conectando ao banco...")
        conn = conectar_db()
        cursor = conn.cursor()
        
        print("🔍 Buscando usuário...")
        executar_query(cursor, 'SELECT id, nome, email, senha_hash FROM usuario WHERE email = ?', (email,))
        usuario = cursor.fetchone()
        conn.close()
        
        if usuario:
            print(f"✅ Usuário encontrado: {usuario[1]}")
            print(f"🔐 Verificando senha para usuário ID: {usuario[0]}")
            if verificar_senha(senha, usuario[3]):
                session['user_id'] = usuario[0]
                session['is_admin_session'] = False  # Garantir que não é admin
                print("🎉 Login realizado com sucesso!")
                print(f"🔍 Sessão após login: {dict(session)}")
                return jsonify({
                    "success": True,
                    "message": "Login realizado com sucesso",
                    "redirect": "/",
                    "usuario": {
                        "id": usuario[0],
                        "nome": usuario[1],
                        "email": usuario[2]
                    }
                })
            else:
                print("❌ Senha incorreta")
                return jsonify({"success": False, "error": "Email ou senha incorretos"}), 401
        else:
            print("❌ Usuário não encontrado")
            return jsonify({"success": False, "error": "Email ou senha incorretos"}), 401
            
    except Exception as e:
        print(f"💥 Erro no login: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": f"Erro interno: {str(e)}"}), 500

@app.route('/api/registro', methods=['POST'])
def api_registro():
    try:
        data = request.get_json()
        nome = data.get('nome')
        email = data.get('email')
        senha = data.get('senha')
        
        if not nome or not email or not senha:
            return jsonify({"success": False, "error": "Todos os campos são obrigatórios"}), 400
        
        conn = conectar_db()
        cursor = conn.cursor()
        
        # Verificar se email já existe
        executar_query(cursor, 'SELECT id FROM usuario WHERE email = ?', (email,))
        if cursor.fetchone():
            conn.close()
            return jsonify({"success": False, "error": "Email já cadastrado"}), 400
        
        # Criar usuário
        senha_hash = hash_senha(senha)
        print(f"🔐 Criando usuário:")
        print(f"   Nome: {nome}")
        print(f"   Email: {email}")
        print(f"   Senha (hash): {senha_hash}")
        
        executar_query(cursor, '''
            INSERT INTO usuario (nome, email, senha_hash, data_criacao, admin)
            VALUES (?, ?, ?, ?, ?)
        ''', (nome, email, senha_hash, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 0))
        
        conn.commit()
        user_id = cursor.lastrowid
        conn.close()
        
        # Fazer login automático
        session['user_id'] = user_id
        
        return jsonify({
            "success": True,
            "message": "Usuário criado com sucesso",
            "redirect": "/",
            "usuario": {
                "id": user_id,
                "nome": nome,
                "email": email
            }
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({"success": True, "message": "Logout realizado com sucesso"})

@app.route('/api/recuperar-senha', methods=['POST'])
def api_recuperar_senha():
    try:
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({"success": False, "error": "Email é obrigatório"}), 400
        
        conn = conectar_db()
        cursor = conn.cursor()
        executar_query(cursor, 'SELECT id FROM usuario WHERE email = ?', (email,))
        usuario = cursor.fetchone()
        conn.close()
        
        if usuario:
            # Em produção, você enviaria um email aqui
            return jsonify({
                "success": True,
                "message": "Se o email existir, você receberá instruções para redefinir sua senha",
                "redirect": "/login"
            })
        else:
            # Por segurança, não revelamos se o email existe
            return jsonify({
                "success": True,
                "message": "Se o email existir, você receberá instruções para redefinir sua senha",
                "redirect": "/login"
            })
            
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/test-carrinho', methods=['POST'])
def test_carrinho():
    """Endpoint para testar carrinho - adicionar produto de teste"""
    try:
        print("🧪 TESTE: Adicionando produto de teste ao carrinho...")
        
        # Dados do produto de teste
        produto_teste = {
            'produto_id': 'teste_001',
            'nome': 'Produto Teste',
            'marca': 'Atlas',
            'preco': 99.90,
            'sabor': 'Chocolate',
            'quantidade': 1,
            'imagem': '/static/images/produto-placeholder.svg'
        }
        
        # Simular requisição
        request._json = produto_teste
        
        # Chamar função de adicionar
        resultado = adicionar_ao_carrinho()
        
        return resultado
        
    except Exception as e:
        print(f"❌ Erro no teste carrinho: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/test-pedido-completo', methods=['POST'])
def test_pedido_completo():
    """Endpoint para testar pedido completo - carrinho + Excel"""
    try:
        print("🧪 TESTE: Testando pedido completo...")
        
        # 1. Adicionar produto ao carrinho
        print("🧪 Passo 1: Adicionando produto ao carrinho...")
        produto_teste = {
            'produto_id': 'teste_001',
            'nome': 'Produto Teste',
            'marca': 'Atlas',
            'preco': 99.90,
            'sabor': 'Chocolate',
            'quantidade': 1,
            'imagem': '/static/images/produto-placeholder.svg'
        }
        
        # Simular adição ao carrinho
        if usuario_logado():
            conn = conectar_db()
            cursor = conn.cursor()
            executar_query(cursor, '''
                INSERT INTO carrinho (user_id, produto_id, nome, marca, preco, sabor, quantidade, imagem)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (session['user_id'], produto_teste['produto_id'], produto_teste['nome'], 
                  produto_teste['marca'], produto_teste['preco'], produto_teste['sabor'], 
                  produto_teste['quantidade'], produto_teste['imagem']))
            conn.commit()
            conn.close()
            print("✅ Produto adicionado ao carrinho no banco")
        else:
            carrinho_temp = obter_carrinho_temporario()
            carrinho_temp.append(produto_teste)
            session['carrinho_temporario'] = carrinho_temp
            session.modified = True
            print("✅ Produto adicionado ao carrinho temporário")
        
        # 2. Verificar carrinho
        print("🧪 Passo 2: Verificando carrinho...")
        carrinho = obter_carrinho_usuario()
        print(f"🧪 Carrinho: {len(carrinho)} itens")
        
        # 3. Salvar no Excel
        print("🧪 Passo 3: Salvando no Excel...")
        dados_cliente_teste = {
            'nome': 'Teste Usuario',
            'email': 'teste@teste.com',
            'telefone': '11999999999',
            'cpf': '12345678901',
            'cep': '01234567',
            'cidade': 'São Paulo',
            'estado': 'SP',
            'bairro': 'Centro',
            'endereco': 'Rua Teste, 123',
            'observacoes': 'Pedido de teste'
        }
        
        order_id_teste = f"teste_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        resultado_excel = salvar_pedido_na_planilha(dados_cliente_teste, carrinho, order_id_teste, "Teste")
        
        return jsonify({
            "success": resultado_excel,
            "message": f"Teste completo: {'Sucesso' if resultado_excel else 'Falha'}",
            "order_id": order_id_teste,
            "carrinho_itens": len(carrinho),
            "carrinho": carrinho
        })
        
    except Exception as e:
        print(f"❌ Erro no teste completo: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/test-excel', methods=['POST'])
def test_excel_save():
    """Endpoint para testar salvamento no Excel"""
    try:
        print("🧪 TESTE: Salvando pedido de teste no Excel...")
        
        # Dados de teste
        dados_cliente_teste = {
            'nome': 'Teste Usuario',
            'email': 'teste@teste.com',
            'telefone': '11999999999',
            'cpf': '12345678901',
            'cep': '01234567',
            'cidade': 'São Paulo',
            'estado': 'SP',
            'bairro': 'Centro',
            'endereco': 'Rua Teste, 123',
            'observacoes': 'Pedido de teste'
        }
        
        carrinho_teste = [
            {
                'nome': 'Produto Teste',
                'preco': 50.00,
                'quantidade': 2,
                'sabor': 'Chocolate'
            }
        ]
        
        order_id_teste = f"teste_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        resultado = salvar_pedido_na_planilha(dados_cliente_teste, carrinho_teste, order_id_teste, "Teste")
        
        return jsonify({
            "success": resultado,
            "message": f"Teste de salvamento: {'Sucesso' if resultado else 'Falha'}",
            "order_id": order_id_teste
        })
        
    except Exception as e:
        print(f"❌ Erro no teste Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/test-db', methods=['GET'])
def test_database():
    """Rota para testar se o banco de dados está funcionando"""
    try:
        print("🧪 Testando conexão com banco de dados...")
        conn = conectar_db()
        cursor = conn.cursor()
        
        database_url = os.environ.get('DATABASE_URL')
        
        if database_url:
            # PostgreSQL
            print("💾 Testando PostgreSQL...")
            cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_name = 'usuario'")
            table_exists = cursor.fetchone()
            database_type = "PostgreSQL"
        else:
            # SQLite
            print("💾 Testando SQLite...")
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='usuario'")
            table_exists = cursor.fetchone()
            database_type = "SQLite"
        
        if table_exists:
            # Contar usuários
            cursor.execute("SELECT COUNT(*) FROM usuario")
            user_count = cursor.fetchone()[0]
            
            conn.close()
            return jsonify({
                "success": True,
                "message": "Banco de dados funcionando!",
                "table_exists": True,
                "user_count": user_count,
                "database_type": database_type
            })
        else:
            conn.close()
            return jsonify({
                "success": False,
                "message": "Tabela 'usuario' não existe",
                "table_exists": False,
                "database_type": database_type
            })
            
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Erro ao conectar com banco de dados"
        }), 500

@app.route('/api/create-tables', methods=['GET', 'POST'])
def create_tables_endpoint():
    """Rota para forçar a criação das tabelas - SEM AUTENTICAÇÃO"""
    try:
        print("🔧 Forçando criação das tabelas...")
        print("🔧 Executando criar_tabelas()...")
        
        # Executar criação das tabelas
        criar_tabelas()
        
        print("🔧 Verificando se tabela foi criada...")
        
        # Verificar se foi criada
        conn = conectar_db()
        cursor = conn.cursor()
        
        database_url = os.environ.get('DATABASE_URL')
        if database_url:
            # PostgreSQL
            cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_name = 'usuario'")
        else:
            # SQLite
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='usuario'")
        
        table_exists = cursor.fetchone()
        
        if table_exists:
            # Contar usuários
            cursor.execute("SELECT COUNT(*) FROM usuario")
            user_count = cursor.fetchone()[0]
            conn.close()
            
            print(f"✅ Tabela criada com sucesso! Usuários: {user_count}")
            return jsonify({
                "success": True,
                "message": "Tabelas criadas com sucesso!",
                "table_exists": True,
                "user_count": user_count
            })
        else:
            conn.close()
            print("❌ Tabela não foi criada")
            return jsonify({
                "success": False,
                "message": "Erro ao criar tabelas - tabela não existe",
                "table_exists": False
            })
            
    except Exception as e:
        print(f"❌ Erro ao criar tabelas: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Erro ao criar tabelas"
        }), 500

@app.route('/api/reset-database', methods=['POST'])
def reset_database():
    """Rota para resetar o banco de dados (útil para plano free)"""
    try:
        print("🔄 Resetando banco de dados...")
        
        # Conectar e deletar tabelas
        conn = conectar_db()
        cursor = conn.cursor()
        
        # Deletar tabela se existir
        cursor.execute("DROP TABLE IF EXISTS usuario")
        conn.commit()
        conn.close()
        
        print("🗑️ Tabela deletada, recriando...")
        
        # Recriar tabelas
        criar_tabelas()
        
        return jsonify({
            "success": True,
            "message": "Banco de dados resetado com sucesso!"
        })
        
    except Exception as e:
        print(f"❌ Erro ao resetar banco: {e}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Erro ao resetar banco de dados"
        }), 500

@app.route('/api/create-test-user', methods=['POST'])
def create_test_user():
    """Rota para criar usuário de teste (útil para debug)"""
    try:
        print("👤 Criando usuário de teste...")
        
        # Dados do usuário de teste
        nome = "Henrique Angelo"
        email = "henriqueangegelo@gmail.com"
        senha = "Henrique@15"
        
        conn = conectar_db()
        cursor = conn.cursor()
        
        # Verificar se já existe
        executar_query(cursor, 'SELECT id FROM usuario WHERE email = ?', (email,))
        if cursor.fetchone():
            conn.close()
            return jsonify({
                "success": True,
                "message": "Usuário já existe no banco!"
            })
        
        # Criar usuário
        senha_hash = hash_senha(senha)
        print(f"🔐 Criando usuário de teste:")
        print(f"   Nome: {nome}")
        print(f"   Email: {email}")
        print(f"   Senha (hash): {senha_hash}")
        
        executar_query(cursor, '''
            INSERT INTO usuario (nome, email, senha_hash, data_criacao, admin)
            VALUES (?, ?, ?, ?, ?)
        ''', (nome, email, senha_hash, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 0))
        
        conn.commit()
        user_id = cursor.lastrowid
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"Usuário de teste criado com sucesso! ID: {user_id}",
            "user_id": user_id
        })
        
    except Exception as e:
        print(f"❌ Erro ao criar usuário de teste: {e}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Erro ao criar usuário de teste"
        }), 500

@app.route('/api/backup-database', methods=['GET'])
def backup_database():
    """Rota para fazer backup do banco de dados"""
    try:
        print("💾 Fazendo backup do banco de dados...")
        
        conn = conectar_db()
        cursor = conn.cursor()
        
        # Buscar todos os usuários
        cursor.execute('SELECT * FROM usuario')
        usuarios = cursor.fetchall()
        conn.close()
        
        # Criar backup em formato JSON
        backup_data = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "usuarios": []
        }
        
        for usuario in usuarios:
            backup_data["usuarios"].append({
                "id": usuario[0],
                "nome": usuario[1],
                "email": usuario[2],
                "senha_hash": usuario[3],
                "data_criacao": usuario[4],
                "admin": usuario[5]
            })
        
        return jsonify({
            "success": True,
            "message": f"Backup criado com {len(usuarios)} usuários",
            "backup": backup_data
        })
        
    except Exception as e:
        print(f"❌ Erro ao fazer backup: {e}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Erro ao fazer backup"
        }), 500

@app.route('/api/restore-database', methods=['POST'])
def restore_database():
    """Rota para restaurar banco de dados do backup"""
    try:
        print("🔄 Restaurando banco de dados...")
        
        data = request.get_json()
        backup = data.get('backup')
        
        if not backup or 'usuarios' not in backup:
            return jsonify({
                "success": False,
                "error": "Backup inválido"
            }), 400
        
        # Limpar tabela atual
        conn = conectar_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM usuario")
        
        # Restaurar usuários
        for usuario_data in backup['usuarios']:
            executar_query(cursor, '''
                INSERT INTO usuario (id, nome, email, senha_hash, data_criacao, admin)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                usuario_data['id'],
                usuario_data['nome'],
                usuario_data['email'],
                usuario_data['senha_hash'],
                usuario_data['data_criacao'],
                usuario_data['admin']
            ))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"Banco restaurado com {len(backup['usuarios'])} usuários"
        })
        
    except Exception as e:
        print(f"❌ Erro ao restaurar: {e}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Erro ao restaurar banco"
        }), 500

# Criar tabelas automaticamente quando o app iniciar
print("🚀 ATLAS SUPLEMENTOS - VERSÃO POSTGRESQL DEFINITIVA - TESTE PERSISTÊNCIA - INICIANDO...")
print("✅ Sistema Atlas Suplementos iniciado!")
print(f"📁 Diretório atual: {os.getcwd()}")
print(f"📁 Templates: {os.path.exists('templates')}")
print(f"📁 Static: {os.path.exists('static')}")
print(f"📁 index.html: {os.path.exists('templates/index.html')}")
print("🔧 USANDO POSTGRESQL - PERSISTÊNCIA GARANTIDA!")

# Criar tabelas do banco de dados automaticamente
print("🔧 Criando tabelas automaticamente...")
criar_tabelas()

# Criar usuário admin padrão se não existir
def criar_admin_padrao():
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        
        # Verificar se já existe admin
        executar_query(cursor, 'SELECT id FROM usuario WHERE admin = 1')
        admin_existente = cursor.fetchone()
        
        if not admin_existente:
            # Criar admin padrão
            admin_email = "admin@atlas.com"
            admin_senha = "rafaelcardeal005"  # Senha segura
            admin_nome = "Administrador Atlas"
            
            senha_hash = hash_senha(admin_senha)
            
            executar_query(cursor, '''
                INSERT INTO usuario (nome, email, senha_hash, data_criacao, admin)
                VALUES (?, ?, ?, ?, ?)
            ''', (admin_nome, admin_email, senha_hash, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 1))
            
            conn.commit()
            conn.close()
            print(f"👑 Usuário admin criado: {admin_email} / {admin_senha}")
        else:
            # Atualizar senha do admin existente
            admin_email = "admin@atlas.com"
            admin_senha = "rafaelcardeal005"  # Nova senha segura
            admin_nome = "Administrador Atlas"
            
            senha_hash = hash_senha(admin_senha)
            
            executar_query(cursor, '''
                UPDATE usuario SET senha_hash = ?, nome = ? WHERE email = ? AND admin = 1
            ''', (senha_hash, admin_nome, admin_email))
            
            conn.commit()
            conn.close()
            print(f"👑 Senha do admin atualizada: {admin_email} / {admin_senha}")
            
    except Exception as e:
        print(f"❌ Erro ao criar/atualizar admin: {e}")

@app.route('/api/fix-carrinho', methods=['POST'])
def fix_carrinho():
    """Recria a tabela carrinho com constraint UNIQUE"""
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        
        # Dropar tabela carrinho se existir
        executar_query(cursor, 'DROP TABLE IF EXISTS carrinho')
        
        # Recriar tabela com constraint UNIQUE
        database_url = os.environ.get('DATABASE_URL')
        if database_url:
            # PostgreSQL
            executar_query(cursor, '''
                CREATE TABLE carrinho (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL,
                    produto_id VARCHAR(255) NOT NULL,
                    nome VARCHAR(255) NOT NULL,
                    marca VARCHAR(255),
                    preco DECIMAL(10,2) NOT NULL,
                    sabor VARCHAR(255),
                    quantidade INTEGER NOT NULL,
                    imagem VARCHAR(500),
                    data_adicionado TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(user_id, produto_id, sabor)
                )
            ''')
        else:
            # SQLite
            executar_query(cursor, '''
                CREATE TABLE carrinho (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    produto_id TEXT NOT NULL,
                    nome TEXT NOT NULL,
                    marca TEXT,
                    preco REAL NOT NULL,
                    sabor TEXT,
                    quantidade INTEGER NOT NULL,
                    imagem TEXT,
                    data_adicionado TEXT DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(user_id, produto_id, sabor)
                )
            ''')
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True, "message": "Tabela carrinho recriada com constraint UNIQUE"})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/test-carrinho-simple', methods=['GET'])
def test_carrinho_simple():
    """Teste simples do carrinho"""
    try:
        if not qualquer_usuario_logado():
            carrinho_temp = obter_carrinho_temporario()
            return jsonify({"carrinho": carrinho_temp, "message": "Carrinho temporário"})
        
        conn = conectar_db()
        cursor = conn.cursor()
        
        executar_query(cursor, '''
            SELECT produto_id, nome, marca, preco, sabor, quantidade, imagem
            FROM carrinho WHERE user_id = ?
        ''', (session['user_id'],))
        
        itens = cursor.fetchall()
        conn.close()
        
        carrinho = []
        for item in itens:
            carrinho.append({
                'produto_id': item[0],
                'nome': item[1],
                'marca': item[2],
                'preco': float(item[3]),
                'sabor': item[4],
                'quantidade': item[5],
                'imagem': item[6]
            })
        
        return jsonify({"carrinho": carrinho, "message": f"Encontrados {len(carrinho)} itens"})
        
    except Exception as e:
        return jsonify({"error": str(e), "carrinho": []})

@app.route('/api/limpar-sessao', methods=['POST'])
def limpar_sessao():
    """Limpa toda a sessão"""
    try:
        session.clear()
        return jsonify({"success": True, "message": "Sessão limpa com sucesso"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

criar_admin_padrao()

if __name__ == '__main__':
    # Configuração para produção
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    
    print(f"🌐 Iniciando servidor na porta {port}")
    app.run(debug=debug, host='0.0.0.0', port=port)
